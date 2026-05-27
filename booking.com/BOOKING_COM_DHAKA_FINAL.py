import sys
sys.stdout.reconfigure(encoding='utf-8')
from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

OUTPUT_FILE = "Dhaka_Hotels_Booking.xlsx"

SEARCH_URL = (
    "https://www.booking.com/searchresults.html"
    "?label=gog235jc-10CAMYyAEoFEIHZ3Vsc2hhbkgzWANobIgBAZgBM7gBF8gBDNgBA"
    "-gBAfgBAYgCAagCAbgCsYTdzgbAAgHSAiQ2ZDFhMDFlZC1hNWYzLTRmYWQtODEwZS1m"
    "NTFiYmM1YWFkZmLYAgHgAgE"
    "&aid=356980&ss=Dhaka%2C+Bangladesh&ssne=Gulshan+1&ssne_untouched=Gulshan+1"
    "&theme_id=58&efdco=1&lang=en-us&dest_id=-2737683&dest_type=city"
    "&ac_position=0&ac_click_type=b&ac_langcode=en&ac_suggestion_list_length=5"
    "&search_selected=true&search_pageview_id=58bf2b18fdaa028e"
    "&checkin=2026-04-10&checkout=2026-04-23"
    "&group_adults=2&no_rooms=1&group_children=0"
    "&sb_travel_purpose=leisure&sb_lp=1&nflt=ht_id%3D204"
    "&chal_t=1775730443440"
    "&force_referer=https%3A%2F%2Fwww.booking.com%2Ffivestars%2Fcity%2Fbd%2F"
    "gulshan.html%3Fchal_t%3D1775714863674%26force_referer%3Dhttps%253A%252F"
    "%252Fwww.google.com%252F"
)

# ─────────────────────────────────────────────
# SCROLL + LOAD MORE (Search Page)
# ─────────────────────────────────────────────
def click_load_more(page):
    while True:
        try:
            btn = page.locator("button:has-text('Load more results')")
            if btn.is_visible():
                print("🔘 Clicking 'Load more results'...")
                btn.scroll_into_view_if_needed()
                btn.click()
                time.sleep(4)
            else:
                break
        except:
            break

def scroll_to_bottom(page):
    last_height, same_count = 0, 0
    while True:
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(2)
        h = page.evaluate("document.body.scrollHeight")
        print(f"🔁 Scroll height: {h}")
        same_count = same_count + 1 if h == last_height else 0
        if same_count > 2:
            break
        last_height = h
    print("✅ Finished scrolling.")

# ─────────────────────────────────────────────
# STEP 1: Extract titles + clean URLs via JS
#         (stays on the search page, no new tabs)
# ─────────────────────────────────────────────
def extract_hotel_list_via_js(page):
    return page.evaluate("""
        () => {
            const cards = document.querySelectorAll("div[data-testid='property-card']");
            return Array.from(cards).map(card => {
                const title = card.querySelector("[data-testid='title']")?.innerText?.trim() || "Unknown";
                let link = card.querySelector("a[data-testid='title-link']")?.getAttribute("href") || "";
                if (link) {
                    if (!link.startsWith("http")) link = "https://www.booking.com" + link;
                    link = link.split("?")[0];
                }
                return { title, link };
            });
        }
    """)

# ─────────────────────────────────────────────
# STEP 2: Fetch hotel page HTML via API request
#         ✅ NO new Chrome tab opened — uses the
#            same browser session cookies silently
# ─────────────────────────────────────────────
def fetch_address_via_api(context, url):
    """
    Uses Playwright context.request.get() — this is a background HTTP request
    that shares the browser's cookies/session but does NOT open a new tab.
    Then parses the HTML with BeautifulSoup to find the address.
    """
    try:
        response = context.request.get(
            url,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                "Accept-Language": "en-US,en;q=0.9",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            },
            timeout=15000
        )

        if not response.ok:
            return f"HTTP {response.status}"

        html = response.text()
        soup = BeautifulSoup(html, "html.parser")

        # ── Selector priority list (hotel detail page) ──
        selectors = [
            {"attrs": {"data-testid": "address"}},
            {"id": "hp_address_subtitle"},
            {"class": "hp_address_subtitle"},
            {"itemprop": "streetAddress"},
            {"class": "address"},
            {"class": "hotel-address"},
        ]

        for sel in selectors:
            el = soup.find(True, sel)
            if el:
                text = el.get_text(separator=" ", strip=True)
                if text and 3 < len(text) < 250:
                    return text

        # ── JSON-LD fallback (address in structured data) ──
        import json, re
        scripts = soup.find_all("script", {"type": "application/ld+json"})
        for s in scripts:
            try:
                data = json.loads(s.string or "")
                # Handle list or single object
                items = data if isinstance(data, list) else [data]
                for item in items:
                    addr = item.get("address", {})
                    if isinstance(addr, dict):
                        parts = [
                            addr.get("streetAddress", ""),
                            addr.get("addressLocality", ""),
                            addr.get("addressCountry", ""),
                        ]
                        combined = ", ".join(p for p in parts if p)
                        if combined:
                            return combined
                    elif isinstance(addr, str) and addr:
                        return addr
            except:
                continue

        return "Not found"

    except Exception as e:
        err = str(e)[:80]
        return f"Error: {err}"

# ─────────────────────────────────────────────
# STEP 3: Save to formatted Excel
# ─────────────────────────────────────────────
def save_to_excel(hotels, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dhaka Hotels"

    HDR_FILL  = PatternFill("solid", fgColor="1F3864")
    HDR_FONT  = Font(name="Calibri", bold=True,   color="FFFFFF", size=11)
    DATA_FONT = Font(name="Calibri", size=10,     color="000000")
    LINK_FONT = Font(name="Calibri", size=10,     color="0563C1", underline="single")
    TITL_FONT = Font(name="Calibri", bold=True,   color="1F3864", size=14)
    META_FONT = Font(name="Calibri", italic=True, color="7F7F7F", size=9)
    FOOT_FONT = Font(name="Calibri", italic=True, color="7F7F7F", size=9)
    THIN      = Side(style="thin", color="D9D9D9")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    C_CTR     = Alignment(horizontal="center", vertical="center")
    C_WRAP    = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)

    ws.merge_cells("A1:D1")
    ws["A1"] = "Dhaka Hotels — Booking.com"
    ws["A1"].font = TITL_FONT
    ws["A1"].alignment = C_CTR
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    ws["A2"] = (
        f"Scraped: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        f"Total Hotels: {len(hotels)}  |  Source: Booking.com"
    )
    ws["A2"].font = META_FONT
    ws["A2"].alignment = C_CTR
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    for col, hdr in enumerate(["#", "Hotel Name", "Address", "Booking.com URL"], 1):
        cell = ws.cell(row=4, column=col, value=hdr)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = C_CTR; cell.border = BORDER
    ws.row_dimensions[4].height = 24

    for i, hotel in enumerate(hotels, 1):
        row = i + 4
        c = ws.cell(row=row, column=1, value=i)
        c.font = DATA_FONT; c.alignment = C_CTR; c.border = BORDER

        c = ws.cell(row=row, column=2, value=hotel["title"])
        c.font = DATA_FONT; c.alignment = C_WRAP; c.border = BORDER

        c = ws.cell(row=row, column=3, value=hotel["address"])
        c.font = DATA_FONT; c.alignment = C_WRAP; c.border = BORDER

        url = hotel["link"]
        c = ws.cell(row=row, column=4, value=url)
        if url:
            c.hyperlink = url
        c.font = LINK_FONT; c.alignment = C_WRAP; c.border = BORDER
        ws.row_dimensions[row].height = 18

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 52

    ws.freeze_panes = "A5"
    last_data_row = len(hotels) + 4
    ws.auto_filter.ref = f"A4:D{last_data_row}"

    foot_row = last_data_row + 3
    ws.merge_cells(f"A{foot_row}:D{foot_row}")
    ws.cell(row=foot_row, column=1,
            value=f"Source: Booking.com  |  Generated: {datetime.now().strftime('%Y-%m-%d')}")
    ws.cell(row=foot_row, column=1).font = FOOT_FONT
    ws.cell(row=foot_row, column=1).alignment = Alignment(horizontal="left")

    wb.save(output_file)
    print(f"\n✅ Excel saved → {output_file}  ({len(hotels)} hotels)")

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def scrape_booking():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=50)
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            )
        )

        # ── Phase 1: Open search page (only tab needed) ──
        page = context.new_page()
        print("🌐 Opening Booking.com search results for Dhaka...")
        page.goto(SEARCH_URL)

        input("\n🧩 Solve CAPTCHA / close popups manually, then press Enter...\n")

        while True:
            scroll_to_bottom(page)
            prev = page.locator("div[data-testid='property-card']").count()
            click_load_more(page)
            scroll_to_bottom(page)
            curr = page.locator("div[data-testid='property-card']").count()
            if curr == prev:
                break

        print(f"\n🏨 Total hotel cards found: {curr}")
        page.evaluate("window.scrollTo(0, 0)")
        time.sleep(1)

        hotel_list = extract_hotel_list_via_js(page)
        print(f"📋 Extracted {len(hotel_list)} hotel URLs\n")

        # ── Phase 2: Fetch address via background HTTP (NO new tabs) ──
        print("📡 Fetching addresses via background requests (no Chrome tabs)...\n")
        hotels_data = []

        for i, hotel in enumerate(hotel_list):
            title = hotel["title"]
            link  = hotel["link"]

            address = fetch_address_via_api(context, link)
            status  = "📍" if address not in ("Not found",) and not address.startswith("Error") else "⚠️"
            print(f"[{i+1:>3}/{len(hotel_list)}] {status} {title}")
            print(f"              {address}")

            hotels_data.append({"title": title, "address": address, "link": link})
            time.sleep(0.8)   # Polite delay between requests

        # ── Phase 3: Save to Excel ──
        save_to_excel(hotels_data, OUTPUT_FILE)

        print("\n🏁 All done. Closing browser.")
        browser.close()

if __name__ == "__main__":
    scrape_booking()
