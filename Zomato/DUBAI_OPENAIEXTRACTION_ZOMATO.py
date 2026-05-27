"""
Enhanced OpenAI GPT-4o-mini Image Text Extraction Script
- Fixed AVIF support with correct package
- Rate limiting protection (12+ second delays)
- Budget monitoring (global only)
- Comprehensive CSV/Excel reporting
- Auto-save, error recovery, and RESUME
- 50 folder batch processing

Updates in this version:
- High-detail vision ("detail":"high") for better price capture
- Stronger extraction prompt to keep prices on the same line
- Smart upscaling for small images (improves legibility)
- REMOVED per-folder budget; process all images until global budget hits
- NEW: Auto-skip completed folders, and resume partial folders only for missing images
"""

import os
import json
import time
import random
from typing import List, Dict, Tuple, Optional, Set
from dataclasses import dataclass, asdict, field
from datetime import datetime
import pandas as pd
from rich.console import Console
from rich.progress import Progress, BarColumn
from rich.table import Table
import base64
import openai
from openai import OpenAI
import threading
from PIL import Image
import io

# Console
console = Console()

# Enhanced image format support
try:
    import pillow_avif  # Correct import for AVIF support
    console.print("[green]✅ AVIF support loaded successfully[/green]")
    AVIF_SUPPORT = True
except ImportError:
    console.print("[yellow]⚠️ AVIF support not available - install pillow-avif-plugin for full compatibility[/yellow]")
    AVIF_SUPPORT = False

# ===================== USER CONFIG =====================
# Set OPENAI_API_KEY in your local environment before running.

SUPPORTED_EXTENSIONS = ['.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.webp', '.gif', '.avif']

# Enhanced rate limiting configuration (UNCHANGED)
RATE_LIMIT_REQUESTS_PER_MINUTE = 35  # Conservative limit
RATE_LIMIT_REQUESTS_PER_HOUR = 1500  # Conservative hourly limit
BASE_SLEEP_BETWEEN_IMAGES = 20       # 12+ seconds minimum
ADAPTIVE_SLEEP_MULTIPLIER = 2
MAX_CONSECUTIVE_FAILURES = 3
MAX_BUDGET = 5.80                  # Global budget ONLY
BUDGET_BUFFER = 0.01

# Set the base path for Dubai menu image data.
BASE_PATH = os.getenv("ZOMATO_MENU_IMAGE_BASE_PATH", "zomato_menu_images")

# Output base
OUTPUT_DIR = os.getenv("ZOMATO_OPENAI_OUTPUT_DIR", "zomato_openai_extracted_results")

# Completion marker filename
COMPLETION_MARK = "_completed.json"

# --- Vision extraction tuning (no impact on rate limiter) ---
UPSCALE_MIN_SIDE = 1200     # upscale small images to at least this longest side
UPSCALE_MAX_FACTOR = 2.0    # never upscale more than 2x

VISION_EXTRACTION_PROMPT = """
Extract menu information in a structured format. For each menu item:
1. Identify the MENU CATEGORY (section headers or main categories)
2. List all MENU ITEMS under each category
3. Include any DESCRIPTION if present
4. Include all PRICES (in AED, DH, DHS or numerical values)

Format rules:
- Start each new category with "CATEGORY:"
- For each item use format:
  ITEM: [item name]
  DESCRIPTION: [description if any]
  PRICE: [price]
- Keep exact spelling and formatting of items
- Maintain original price format
- Separate items with a line break

Example format:
CATEGORY: WHITE
ITEM: NV Perrier-Jouët, Grand Brut
PRICE: 870 AED

ITEM: NV Bollinger, Special Cuvée
PRICE: 990 AED
"""

# ===================== DATA CLASSES =====================
@dataclass
class ImageRecord:
    image_path: str
    folder: str
    extracted_text: Optional[str] = None
    confidence: Optional[float] = None
    processing_time: float = 0.0
    timestamp: str = ""
    error: Optional[str] = None
    model_used: str = "gpt-4o-mini"
    prompt_tokens: int = 0
    completion_tokens: int = 0
    cost_usd: float = 0.0
    image_size_bytes: int = 0
    detected_format: str = ""

@dataclass
class TokenTracker:
    requests_per_minute: int = RATE_LIMIT_REQUESTS_PER_MINUTE
    requests_per_hour: int = RATE_LIMIT_REQUESTS_PER_HOUR
    minute_requests: List[float] = field(default_factory=list)
    hour_requests: List[float] = field(default_factory=list)
    total_tokens_used: int = 0
    total_cost: float = 0.0

    def add_request(self, tokens_used: int = 0, cost: float = 0.0):
        now = time.time()
        self.minute_requests.append(now)
        self.hour_requests.append(now)
        self.total_tokens_used += tokens_used
        self.total_cost += cost
        self._clean_old_requests()

    def _clean_old_requests(self):
        now = time.time()
        minute_ago = now - 60
        hour_ago = now - 3600
        self.minute_requests = [req for req in self.minute_requests if req > minute_ago]
        self.hour_requests = [req for req in self.hour_requests if req > hour_ago]

    def can_make_request(self) -> Tuple[bool, str]:
        self._clean_old_requests()
        if len(self.minute_requests) >= self.requests_per_minute:
            return False, f"Rate limit: {len(self.minute_requests)}/{self.requests_per_minute} requests in last minute"
        if len(self.hour_requests) >= self.requests_per_hour:
            return False, f"Rate limit: {len(self.hour_requests)}/{self.requests_per_hour} requests in last hour"
        if self.total_cost >= MAX_BUDGET - BUDGET_BUFFER:
            return False, f"Budget limit: ${self.total_cost:.3f}/${MAX_BUDGET}"
        return True, "OK"

    def can_process_image(self) -> Tuple[bool, str]:
        return self.can_make_request()

    def get_status(self) -> Dict:
        self._clean_old_requests()
        return {
            "requests_last_minute": len(self.minute_requests),
            "requests_last_hour": len(self.hour_requests),
            "total_tokens": self.total_tokens_used,
            "total_cost": self.total_cost,
            "remaining_budget": MAX_BUDGET - self.total_cost
        }

@dataclass
class RateLimiter:
    token_tracker: TokenTracker = field(default_factory=TokenTracker)
    _lock: threading.Lock = field(default_factory=threading.Lock)

    def wait_if_needed(self):
        with self._lock:
            can_proceed, reason = self.token_tracker.can_make_request()
            if not can_proceed:
                if "Rate limit" in reason:
                    wait_time = 65
                    console.print(f"[yellow]⏳ {reason}. Waiting {wait_time}s...[/yellow]")
                    time.sleep(wait_time)
                elif "Budget limit" in reason:
                    console.print(f"[red]💰 {reason}. Stopping processing.[/red]")
                    raise Exception(f"Budget exhausted: {reason}")

@dataclass
class EnhancedBatchProgress:
    total_folders: int
    processed_folders: int = 0
    total_images: int = 0
    processed_images: int = 0
    successful_images: int = 0
    failed_images: int = 0
    total_cost: float = 0.0
    total_tokens_used: int = 0
    folder_costs: Dict[str, float] = field(default_factory=dict)
    folder_tokens: Dict[str, int] = field(default_factory=dict)
    folder_image_counts: Dict[str, int] = field(default_factory=dict)
    failed_folders: List[str] = field(default_factory=list)
    start_time: float = field(default_factory=time.time)

# ===================== HELPERS =====================
def adaptive_sleep_between_images(consecutive_failures: int = 0, folder_progress: float = 0.0):
    base_sleep = BASE_SLEEP_BETWEEN_IMAGES
    failure_multiplier = 1 + (consecutive_failures * 0.5)
    progress_multiplier = 1 + (folder_progress * 0.3)
    jitter = random.uniform(0.8, 1.2)
    total_sleep = base_sleep * failure_multiplier * progress_multiplier * jitter
    console.print(f"[dim]😴 Sleeping {total_sleep:.1f}s (base: {base_sleep}s, failures: {consecutive_failures}, progress: {folder_progress:.1f})[/dim]")
    time.sleep(total_sleep)

def detect_image_format_and_convert(image_path: str) -> Tuple[str, str]:
    """
    Detect image format and convert to a high-quality JPEG.
    Also upscale small images (helps pick up faint/right-aligned prices).
    Returns: (base64_string, detected_format)
    """
    try:
        with Image.open(image_path) as img:
            actual_format = img.format.lower() if img.format else None

            # Normalize to RGB
            if img.mode in ('RGBA', 'LA', 'P'):
                rgb = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                rgb.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                img = rgb
            elif img.mode != 'RGB':
                img = img.convert('RGB')

            # Gentle upscale for small images
            w, h = img.size
            longest = max(w, h)
            if longest < UPSCALE_MIN_SIDE:
                scale = min(UPSCALE_MIN_SIDE / float(longest), UPSCALE_MAX_FACTOR)
                new_size = (int(w * scale), int(h * scale))
                img = img.resize(new_size, Image.LANCZOS)

            # Save to high-quality JPEG in-memory
            buf = io.BytesIO()
            img.save(buf, format='JPEG', quality=95, optimize=True)
            buf.seek(0)

            base64_string = base64.b64encode(buf.getvalue()).decode('utf-8')
            console.print(f"[dim]🖼️ Converted {actual_format or 'unknown'} → JPEG "
                          f"({img.size[0]}x{img.size[1]}) for {os.path.basename(image_path)}[/dim]")
            return base64_string, actual_format or "jpeg"

    except Exception as e:
        console.print(f"[yellow]⚠️ PIL conversion failed for {os.path.basename(image_path)}: {e}[/yellow]")
        try:
            with open(image_path, "rb") as image_file:
                base64_string = base64.b64encode(image_file.read()).decode('utf-8')
                return base64_string, "unknown"
        except Exception as e2:
            raise Exception(f"Both PIL and direct encoding failed: {e2}")

def get_image_size(image_path: str) -> int:
    return os.path.getsize(image_path)

def calculate_openai_cost(prompt_tokens: int, completion_tokens: int, model: str = "gpt-4o-mini") -> float:
    if model == "gpt-4o-mini":
        # GPT-4o-mini pricing: $0.15 per 1M input tokens, $0.60 per 1M output tokens
        input_cost = (prompt_tokens / 1_000_000) * 0.15
        output_cost = (completion_tokens / 1_000_000) * 0.60
        return input_cost + output_cost
    return 0.0

def list_images_with_enhanced_validation(folder_path: str) -> List[str]:
    images = []
    skipped_details = []
    format_stats = {}

    if not os.path.isdir(folder_path):
        return images

    for file in os.listdir(folder_path):
        if any(file.lower().endswith(ext) for ext in SUPPORTED_EXTENSIONS):
            full_path = os.path.join(folder_path, file)

            # Check file size
            try:
                file_size = os.path.getsize(full_path)
                if file_size == 0:
                    skipped_details.append(f"{file}: Zero-byte file")
                    continue
                elif file_size < 100:
                    skipped_details.append(f"{file}: File too small ({file_size} bytes)")
                    continue
            except OSError as e:
                skipped_details.append(f"{file}: File access error - {e}")
                continue

            # Enhanced format detection and validation
            try:
                actual_format = "unknown"
                with open(full_path, 'rb') as f:
                    header = f.read(20)

                if header.startswith(b'\x89PNG'):
                    actual_format = "PNG"
                elif header.startswith(b'\xff\xd8\xff'):
                    actual_format = "JPEG"
                elif header.startswith(b'RIFF') and b'WEBP' in header:
                    actual_format = "WebP"
                elif header[4:12] == b'ftypavif':
                    actual_format = "AVIF"
                    if not AVIF_SUPPORT:
                        skipped_details.append(f"{file}: AVIF file detected but pillow-avif-plugin not installed")
                        continue
                elif header.startswith(b'BM'):
                    actual_format = "BMP"
                elif header.startswith(b'GIF87a') or header.startswith(b'GIF89a'):
                    actual_format = "GIF"

                with Image.open(full_path) as img:
                    img.verify()
                with Image.open(full_path) as img:
                    if img.size[0] < 10 or img.size[1] < 10:
                        skipped_details.append(f"{file}: Image too small ({img.size})")
                        continue
                    format_stats[actual_format] = format_stats.get(actual_format, 0) + 1
                    images.append(full_path)

            except Exception as e:
                error_msg = str(e)
                if "AVIF support not installed" in error_msg:
                    skipped_details.append(f"{file}: AVIF format requires pillow-avif-plugin")
                elif "cannot identify image file" in error_msg:
                    skipped_details.append(f"{file}: Unrecognized format ({actual_format})")
                else:
                    skipped_details.append(f"{file}: PIL error - {error_msg[:50]}...")

    if skipped_details or format_stats:
        folder_name = os.path.basename(folder_path)
        if format_stats:
            format_summary = ", ".join([f"{fmt}: {count}" for fmt, count in format_stats.items()])
            console.print(f"[green]📊 {folder_name} - Found formats: {format_summary}[/green]")
        if skipped_details:
            console.print(f"[yellow]⚠️ {folder_name} - Skipped {len(skipped_details)} files:[/yellow]")
            error_groups = {}
            for detail in skipped_details:
                error_type = detail.split(": ")[-1].split(" ")[0]
                if error_type not in error_groups:
                    error_groups[error_type] = []
                error_groups[error_type].append(detail)
            for error_type, errors in error_groups.items():
                console.print(f"[dim]  - {error_type}: {len(errors)} files[/dim]")
                if len(errors) <= 3:
                    for error in errors:
                        console.print(f"[dim]    • {error}[/dim]")
                else:
                    for error in errors[:2]:
                        console.print(f"[dim]    • {error}[/dim]")
                    console.print(f"[dim]    • ... and {len(errors) - 2} more[/dim]")

    return sorted(images)

def safe_folder_key(folder_path: str) -> str:
    folder_name = os.path.basename(folder_path.rstrip(os.sep))
    safe_name = "".join(c for c in folder_name if c.isalnum() or c in ('-', '_', ' '))
    return safe_name or "unknown_folder"

def ensure_dirs(base_dir: str, folder_key: str) -> str:
    out_dir = os.path.join(base_dir, folder_key)
    os.makedirs(out_dir, exist_ok=True)
    return out_dir

def folder_is_complete(folder_path: str, out_dir: str, imgs: List[str]) -> bool:
    """Return True if every image has a corresponding .txt output (fast skip)."""
    image_bases: Set[str] = {os.path.splitext(os.path.basename(p))[0] for p in imgs}
    existing_txt: Set[str] = {
        os.path.splitext(fn)[0]
        for fn in os.listdir(out_dir)
        if fn.lower().endswith(".txt")
    }
    # All images done?
    if image_bases.issubset(existing_txt) and len(existing_txt) >= len(image_bases):
        # If marker exists, great; if not, create it now
        mark_path = os.path.join(out_dir, COMPLETION_MARK)
        if not os.path.exists(mark_path):
            with open(mark_path, "w", encoding="utf-8") as fm:
                json.dump({
                    "completed_at": datetime.now().isoformat(),
                    "images": len(image_bases),
                    "note": "Auto-marked complete on resume check"
                }, fm, ensure_ascii=False, indent=2)
        return True
    # If a marker exists, we still validate against images to avoid stale markers
    return False

def write_completion_marker(out_dir: str, imgs_count: int, added_tokens: int, added_cost: float):
    mark_path = os.path.join(out_dir, COMPLETION_MARK)
    with open(mark_path, "w", encoding="utf-8") as fm:
        json.dump({
            "completed_at": datetime.now().isoformat(),
            "images": imgs_count,
            "added_tokens_this_run": added_tokens,
            "added_cost_this_run": added_cost
        }, fm, ensure_ascii=False, indent=2)

def ocr_one_image_openai_only(image_path: str, model_name: str, rate_limiter: RateLimiter) -> ImageRecord:
    start_time = time.time()
    folder = os.path.dirname(image_path)

    record = ImageRecord(
        image_path=image_path,
        folder=folder,
        timestamp=datetime.now().isoformat(),
        model_used=model_name,
        image_size_bytes=get_image_size(image_path)
    )

    try:
        # Rate limiting check (UNCHANGED)
        rate_limiter.wait_if_needed()

        # Enhanced image encoding with format detection (includes upscaling)
        base64_image, detected_format = detect_image_format_and_convert(image_path)
        record.detected_format = detected_format

        # Create OpenAI client
        client = OpenAI()

        # High-detail + stronger prompt, more max_tokens
        response = client.chat.completions.create(
            model=model_name,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": VISION_EXTRACTION_PROMPT},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/jpeg;base64,{base64_image}",
                                "detail": "high"
                            }
                        }
                    ]
                }
            ],
            max_tokens=1500,
            temperature=0
        )

        # Extract response data
        if response.choices and response.choices[0].message:
            record.extracted_text = response.choices[0].message.content
            record.confidence = 1.0

        # Token usage and cost calculation
        if hasattr(response, 'usage') and response.usage:
            record.prompt_tokens = response.usage.prompt_tokens
            record.completion_tokens = response.usage.completion_tokens
            record.cost_usd = calculate_openai_cost(
                record.prompt_tokens,
                record.completion_tokens,
                model_name
            )

        # Update rate limiter (global budget accounting)
        total_tokens = record.prompt_tokens + record.completion_tokens
        rate_limiter.token_tracker.add_request(total_tokens, record.cost_usd)

        # Optional sanity warning if no AED noticed
        if record.extracted_text and (" AED" not in record.extracted_text):
            console.print(f"[yellow]⚠️ No 'AED' detected in {os.path.basename(image_path)} output. "
                          f"If this image has prices, the text may be too faint or cropped.[/yellow]")

        console.print(f"[green]✅ {os.path.basename(image_path)}: ${record.cost_usd:.4f}, {total_tokens} tokens[/green]")

    except openai.RateLimitError as e:
        record.error = f"Rate limit error: {str(e)}"
        console.print(f"[red]🚫 Rate limit hit for {os.path.basename(image_path)}: {record.error}[/red]")
        time.sleep(60)

    except openai.APIError as e:
        record.error = f"API error: {str(e)}"
        console.print(f"[red]❌ API error for {os.path.basename(image_path)}: {record.error}[/red]")

    except Exception as e:
        record.error = f"Unexpected error: {str(e)}"
        console.print(f"[red]💥 Unexpected error for {os.path.basename(image_path)}: {record.error}[/red]")

    finally:
        record.processing_time = time.time() - start_time

    return record

def parse_structured_menu_text(text: str) -> List[Dict]:
    """Parse extracted text into structured menu items."""
    items = []
    current_category = ""
    current_item = {}
    
    for line in text.split('\n'):
        line = line.strip()
        if not line:
            if current_item:
                current_item['category'] = current_category
                items.append(current_item)
                current_item = {}
            continue
            
        if line.startswith('CATEGORY:'):
            if current_item:
                current_item['category'] = current_category
                items.append(current_item)
                current_item = {}
            current_category = line.replace('CATEGORY:', '').strip()
            
        elif line.startswith('ITEM:'):
            if current_item:
                current_item['category'] = current_category
                items.append(current_item)
            current_item = {'menu_item': line.replace('ITEM:', '').strip()}
            
        elif line.startswith('DESCRIPTION:'):
            if current_item:
                current_item['description'] = line.replace('DESCRIPTION:', '').strip()
                
        elif line.startswith('PRICE:'):
            if current_item:
                current_item['price'] = line.replace('PRICE:', '').strip()
    
    # Add last item if exists
    if current_item:
        current_item['category'] = current_category
        items.append(current_item)
        
    return items

# Update the convert_txt_outputs_to_structured_formats function
def convert_txt_outputs_to_structured_formats(out_dir: str, records: List[ImageRecord]):
    if not records:
        return

    structured_data = []
    for rec in records:
        if rec.extracted_text:
            menu_items = parse_structured_menu_text(rec.extracted_text)
            for item in menu_items:
                structured_data.append({
                    'image_file': os.path.basename(rec.image_path),
                    'folder': os.path.basename(rec.folder),
                    'menu_category': item.get('category', ''),
                    'menu_item': item.get('menu_item', ''),
                    'description': item.get('description', ''),
                    'price': item.get('price', ''),
                    'processing_time': rec.processing_time,
                    'timestamp': rec.timestamp,
                    'model_used': rec.model_used,
                    'success': "Yes" if not rec.error else "No"
                })

    # Create DataFrame with structured columns
    df = pd.DataFrame(structured_data)
    
    # Reorder columns for better readability
    column_order = [
        'menu_category',
        'menu_item',
        'description',
        'price',
        'image_file',
        'folder',
        'processing_time',
        'timestamp',
        'model_used',
        'success'
    ]
    df = df[column_order]

    # Export to CSV and Excel
    csv_path = os.path.join(out_dir, "structured_menu_data.csv")
    xlsx_path = os.path.join(out_dir, "structured_menu_data.xlsx")
    
    df.to_csv(csv_path, index=False, encoding='utf-8')
    try:
        df.to_excel(xlsx_path, index=False, engine='openpyxl')
        # Add Excel formatting
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font
        
        wb = load_workbook(xlsx_path)
        ws = wb.active
        
        # Style header row
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
        wb.save(xlsx_path)
        
    except ImportError:
        console.print("[yellow]⚠️ openpyxl not installed, skipping Excel export[/yellow]")

    console.print(f"[green]💾 Structured menu data saved: {csv_path}[/green]")

def generate_comprehensive_summary_report(progress: EnhancedBatchProgress, output_dir: str, token_tracker: TokenTracker):
    folder_summary_data = []
    total_elapsed = time.time() - progress.start_time

    for folder_name in progress.folder_costs.keys():
        folder_summary_data.append({
            'folder_name': folder_name,
            'images_processed': progress.folder_image_counts.get(folder_name, 0),
            'tokens_used': progress.folder_tokens.get(folder_name, 0),
            'cost_usd': progress.folder_costs.get(folder_name, 0.0),
            'avg_cost_per_image': progress.folder_costs.get(folder_name, 0.0) / max(1, progress.folder_image_counts.get(folder_name, 1)),
            'status': 'Failed' if folder_name in progress.failed_folders else 'Success'
        })

    overall_summary = {
        'total_folders_attempted': len(progress.folder_costs),
        'successful_folders': len(progress.folder_costs) - len(progress.failed_folders),
        'failed_folders': len(progress.failed_folders),
        'total_images_processed': progress.processed_images,
        'successful_images': progress.successful_images,
        'failed_images': progress.failed_images,
        'success_rate_percent': (progress.successful_images / max(1, progress.processed_images)) * 100,
        'total_tokens_used': progress.total_tokens_used,
        'total_cost_usd': progress.total_cost,
        'avg_cost_per_image': progress.total_cost / max(1, progress.successful_images),
        'processing_time_minutes': total_elapsed / 60,
        'images_per_minute': progress.processed_images / max(1, total_elapsed / 60),
        'remaining_budget': MAX_BUDGET - progress.total_cost,
        'estimated_images_remaining': int((MAX_BUDGET - progress.total_cost) / max(0.001, progress.total_cost / max(1, progress.successful_images)))
    }

    if folder_summary_data:
        folder_df = pd.DataFrame(folder_summary_data)
        folder_csv_path = os.path.join(output_dir, "folder_summary_report.csv")
        folder_xlsx_path = os.path.join(output_dir, "folder_summary_report.xlsx")
        folder_df.to_csv(folder_csv_path, index=False)
        try:
            folder_df.to_excel(folder_xlsx_path, index=False, engine='openpyxl')
        except ImportError:
            pass
        console.print(f"[green]📊 Folder summary saved: {folder_csv_path}[/green]")

    overall_df = pd.DataFrame([overall_summary])
    overall_csv_path = os.path.join(output_dir, "overall_batch_summary.csv")
    overall_xlsx_path = os.path.join(output_dir, "overall_batch_summary.xlsx")
    overall_df.to_csv(overall_csv_path, index=False)
    try:
        overall_df.to_excel(overall_xlsx_path, index=False, engine='openpyxl')
    except ImportError:
        pass

    console.print(f"[green]📊 Overall summary saved: {overall_csv_path}[/green]")
    return overall_summary

def get_all_restaurant_folders(base_path: str) -> List[str]:
    if not os.path.isdir(base_path):
        console.print(f"[red]❌ Base path does not exist: {base_path}[/red]")
        return []
    folders = []
    for item in os.listdir(base_path):
        item_path = os.path.join(base_path, item)
        if os.path.isdir(item_path):
            if list_images_with_enhanced_validation(item_path):
                folders.append(item_path)
    return sorted(folders)

def get_limited_restaurant_folders(base_path: str, max_folders: int = 50, skip_completed: bool = True) -> List[str]:
    """Return up to N folders; optionally skip those already fully completed."""
    all_folders = get_all_restaurant_folders(base_path)
    all_folders.sort()
    selected = []
    skipped = 0

    for folder_path in all_folders:
        if len(selected) >= max_folders:
            break
        # Quick pre-filter: if out_dir has completion mark and matches image count, skip
        folder_key = safe_folder_key(folder_path)
        out_dir = os.path.join(OUTPUT_DIR, folder_key)
        imgs = list_images_with_enhanced_validation(folder_path)
        if skip_completed and os.path.isdir(out_dir):
            if folder_is_complete(folder_path, out_dir, imgs):
                skipped += 1
                continue
        if imgs:
            selected.append(folder_path)

    console.print(f"[cyan]📁 Selected {len(selected)} folders out of {len(all_folders)} total ({skipped} auto-skipped as complete)[/cyan]")
    return selected

def bulletproof_process_single_folder(folder_path: str, model_name: str, rate_limiter: RateLimiter,
                                      overall_progress: EnhancedBatchProgress) -> Tuple[bool, float, int]:
    folder_name = os.path.basename(folder_path)
    imgs = list_images_with_enhanced_validation(folder_path)

    if not imgs:
        console.print(f"[yellow]📁 {folder_name}: No valid images found[/yellow]")
        return True, 0.0, 0

    folder_key = safe_folder_key(folder_path)
    out_dir = ensure_dirs(OUTPUT_DIR, folder_key)
    jsonl_path = os.path.join(out_dir, "ocr_results.jsonl")

    # Resume-friendly: append if exists
    existing_txt = {
        os.path.splitext(fn)[0]
        for fn in os.listdir(out_dir)
        if fn.lower().endswith(".txt")
    }
    jsonl_mode = "a" if os.path.exists(jsonl_path) else "w"

    # If folder already fully complete, skip it entirely
    if folder_is_complete(folder_path, out_dir, imgs):
        console.print(f"[dim]✅ {folder_name}: Already complete — skipping[/dim]")
        return True, 0.0, 0

    # Track totals for this run
    folder_totals = {"images": len(imgs), "cost": 0.0, "failures": 0, "successes": 0, "tokens": 0, "consecutive_failures": 0}
    all_records: List[ImageRecord] = []

    console.print(f"[bold cyan]🍽️ Processing {folder_name} ({len(imgs)} images) - Global budget only[/bold cyan]")

    try:
        with Progress(
            "[progress.description]{task.description}",
            BarColumn(),
            "[progress.percentage]{task.percentage:>3.0f}%",
            console=console,
        ) as progress_bar, open(jsonl_path, jsonl_mode, encoding="utf-8") as fjl:

            task = progress_bar.add_task(f"[green]{folder_name}...", total=len(imgs))

            for img_idx, img_path in enumerate(imgs):
                # Skip already done images (resume)
                base = os.path.splitext(os.path.basename(img_path))[0]
                if base in existing_txt:
                    progress_bar.advance(task)
                    continue

                # Global budget check (kept)
                if overall_progress.total_cost >= MAX_BUDGET - BUDGET_BUFFER:
                    console.print(f"[red]💰 Global budget nearly exhausted, stopping[/red]")
                    return False, folder_totals["cost"], len(all_records)

                # Failure threshold check
                if folder_totals["consecutive_failures"] >= MAX_CONSECUTIVE_FAILURES:
                    console.print(f"[red]❌ Too many consecutive failures ({folder_totals['consecutive_failures']}), skipping folder[/red]")
                    overall_progress.failed_folders.append(folder_name)
                    return False, folder_totals["cost"], len(all_records)

                can_process, reason = rate_limiter.token_tracker.can_process_image()
                if not can_process:
                    console.print(f"[red]🛑 Token/budget gate: {reason}[/red]")
                    return False, folder_totals["cost"], len(all_records)

                try:
                    rec = ocr_one_image_openai_only(img_path, model_name, rate_limiter)
                    all_records.append(rec)

                    txt_name = base + ".txt"
                    with open(os.path.join(out_dir, txt_name), "w", encoding="utf-8") as ft:
                        ft.write(rec.extracted_text or "")

                    fjl.write(json.dumps(asdict(rec), ensure_ascii=False) + "\n")
                    fjl.flush()

                    folder_totals["cost"] += rec.cost_usd
                    folder_totals["tokens"] += (rec.prompt_tokens + rec.completion_tokens)
                    overall_progress.total_cost += rec.cost_usd
                    overall_progress.total_tokens_used += (rec.prompt_tokens + rec.completion_tokens)
                    overall_progress.processed_images += 1

                    if rec.error:
                        folder_totals["failures"] += 1
                        folder_totals["consecutive_failures"] += 1
                        overall_progress.failed_images += 1
                        console.print(f"[yellow]⚠️ Image {img_idx+1}/{len(imgs)} failed: {rec.error}[/yellow]")
                    else:
                        folder_totals["successes"] += 1
                        folder_totals["consecutive_failures"] = 0
                        overall_progress.successful_images += 1

                    folder_progress = (img_idx + 1) / len(imgs)
                    adaptive_sleep_between_images(folder_totals["consecutive_failures"], folder_progress)

                except Exception as e:
                    console.print(f"[red]❌ Critical error processing image {img_idx+1}: {e}[/red]")
                    folder_totals["consecutive_failures"] += 1
                    overall_progress.failed_images += 1
                    time.sleep(30)

                progress_bar.advance(task)

    except KeyboardInterrupt:
        console.print(f"\n[yellow]⏸️ {folder_name}: Interrupted by user[/yellow]")
        overall_progress.failed_folders.append(folder_name)
        return False, folder_totals["cost"], len(all_records)

    # At the end, if everything is done (all images now have txt), drop a completion marker
    imgs_bases = {os.path.splitext(os.path.basename(p))[0] for p in imgs}
    now_done_txt = {
        os.path.splitext(fn)[0]
        for fn in os.listdir(out_dir)
        if fn.lower().endswith(".txt")
    }
    if imgs_bases.issubset(now_done_txt):
        write_completion_marker(out_dir, len(imgs_bases), folder_totals["tokens"], folder_totals["cost"])

    overall_progress.folder_costs[folder_name] = folder_totals["cost"]
    overall_progress.folder_tokens[folder_name] = folder_totals["tokens"]
    overall_progress.folder_image_counts[folder_name] = len(now_done_txt)

    if all_records:
        convert_txt_outputs_to_structured_formats(out_dir, all_records)

    success_rate = (folder_totals["successes"] / max(1, len(all_records))) * 100
    console.print(
        f"[green]✅ {folder_name}: processed {len(all_records)} images this run "
        f"({folder_totals['successes']} successes, {success_rate:.1f}%), "
        f"{folder_totals['tokens']:,} tokens, ${folder_totals['cost']:.4f} added[/green]"
    )

    return True, folder_totals['cost'], len(all_records)

def display_progress_table(progress: EnhancedBatchProgress, token_tracker: TokenTracker):
    table = Table(title="Batch Processing Progress")
    table.add_column("Metric", justify="left", style="cyan")
    table.add_column("Value", justify="right", style="green")

    elapsed_time = time.time() - progress.start_time
    success_rate = (progress.successful_images / max(1, progress.processed_images)) * 100

    table.add_row("Folders Processed", f"{progress.processed_folders}/{progress.total_folders}")
    table.add_row("Images Processed", f"{progress.processed_images}")
    table.add_row("Successful Images", f"{progress.successful_images}")
    table.add_row("Failed Images", f"{progress.failed_images}")
    table.add_row("Success Rate", f"{success_rate:.1f}%")
    table.add_row("Total Tokens", f"{progress.total_tokens_used:,}")
    table.add_row("Total Cost", f"${progress.total_cost:.4f}")
    table.add_row("Remaining Budget", f"${MAX_BUDGET - progress.total_cost:.4f}")
    hours = elapsed_time / 3600
    minutes = (elapsed_time % 3600) / 60
    table.add_row("Processing Time", f"{int(hours)}h {minutes:.1f}m")
    table.add_row("Images/Minute", f"{progress.processed_images / max(1, elapsed_time/60):.1f}")
    console.print(table)

# ===================== MAIN =====================
def main():
    console.print(f"[cyan]🖼️ Image format support:[/cyan]")
    console.print(f"[cyan]  - JPEG/PNG/WebP/BMP/GIF: ✅ Built-in[/cyan]")
    console.print(f"[cyan]  - AVIF: {'✅ Available' if AVIF_SUPPORT else '❌ Install pillow-avif-plugin'}[/cyan]")

    if os.environ.get("OPENAI_API_KEY", "").strip() in ("", "paste-your-api-key-here"):
        console.print("[red]❌ Please set OPENAI_API_KEY in your environment before running.[/red]")
        return 1

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    console.print(f"[cyan]📁 Processing folders from: {BASE_PATH}[/cyan]")

    # Build folder list with pre-skip of completed folders
    all_folders = get_limited_restaurant_folders(BASE_PATH, max_folders=50, skip_completed=True)
    if not all_folders:
        console.print("[yellow]ℹ️ Nothing to do — all folders appear complete (or no valid images).[/yellow]")
        return 0

    overall_progress = EnhancedBatchProgress(total_folders=len(all_folders))
    rate_limiter = RateLimiter()

    console.print(f"[bold green]🚀 Starting batch processing of {len(all_folders)} folders[/bold green]")
    console.print(f"[cyan]💰 Global Budget: ${MAX_BUDGET:.2f} (no per-folder cap)[/cyan]")
    console.print(f"[cyan]⏱️ Sleep between images: {BASE_SLEEP_BETWEEN_IMAGES}+ seconds[/cyan]")
    console.print(f"[cyan]🖼️ Upscaling small images + high-detail vision enabled[/cyan]")

    try:
        for folder_idx, folder_path in enumerate(all_folders):
            console.print(f"\n[bold blue]📁 Folder {folder_idx + 1}/{len(all_folders)}: {os.path.basename(folder_path)}[/bold blue]")

            success, folder_cost, images_processed = bulletproof_process_single_folder(
                folder_path, "gpt-4o-mini", rate_limiter, overall_progress
            )

            overall_progress.processed_folders += 1
            display_progress_table(overall_progress, rate_limiter.token_tracker)

            if overall_progress.total_cost >= MAX_BUDGET - BUDGET_BUFFER:
                console.print(f"[red]💰 Global budget exhausted (${overall_progress.total_cost:.4f}/${MAX_BUDGET}), stopping[/red]")
                break

            if folder_idx < len(all_folders) - 1:
                console.print("[dim]⏸️ Sleeping 32s between folders...[/dim]")
                time.sleep(32)

    except KeyboardInterrupt:
        console.print("\n[yellow]⏸️ Batch processing interrupted by user[/yellow]")

    except Exception as e:
        console.print(f"\n[red]💥 Critical error during batch processing: {e}[/red]")

    finally:
        console.print("\n[bold cyan]📊 Generating summary reports...[/bold cyan]")
        summary = generate_comprehensive_summary_report(
            overall_progress, OUTPUT_DIR, rate_limiter.token_tracker
        )

        console.print("\n[bold green]🎉 Batch Processing Complete![/bold green]")
        console.print(f"[green]✅ Successfully processed {summary['successful_images']} images[/green]")
        console.print(f"[yellow]⚠️ Failed: {summary['failed_images']} images[/yellow]")
        console.print(f"[cyan]💰 Total cost: ${summary['total_cost_usd']:.4f}[/cyan]")
        console.print(f"[cyan]🏦 Remaining budget: ${summary['remaining_budget']:.4f}[/cyan]")
        hours = summary['processing_time_minutes'] / 60
        minutes = (summary['processing_time_minutes'] % 60)
        console.print(f"[cyan]⏱️ Processing time: {int(hours)}h {minutes:.1f}m[/cyan]")
        if summary['estimated_images_remaining'] > 0:
            console.print(f"[blue]📈 Estimated images remaining with current budget: {summary['estimated_images_remaining']}[/blue]")

    return 0

if __name__ == "__main__":
    exit(main())
