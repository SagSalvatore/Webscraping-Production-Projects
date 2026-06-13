from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = r"C:\Users\SagarSingh\OneDrive - Mordor Intelligence\Desktop\Webscraping-Production-Projects\talabat\data\urls\entity_resolution_validation.xlsx"

wb = load_workbook(PATH)

# ── helpers ───────────────────────────────────────────────────────────────
thin = Side(style="thin", color="CCCCCC")
def bdr(): return Border(left=thin, right=thin, top=thin, bottom=thin)

def s(ws, row, col, value, bold=False, bg=None, fg="000000",
      align="left", wrap=False, italic=False, size=10, formula_color=False):
    c = ws.cell(row=row, column=col, value=value)
    color = fg if not formula_color else "000000"
    c.font = Font(name="Arial", bold=bold, color=color, italic=italic, size=size)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = bdr()
    return c

# ── create sheet ──────────────────────────────────────────────────────────
ws = wb.create_sheet("Distance Formula Guide")

# column widths
widths = {"A": 28, "B": 22, "C": 22, "D": 20, "E": 20, "F": 18, "G": 38}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# ── TITLE ─────────────────────────────────────────────────────────────────
ws.merge_cells("A1:G1")
c = ws.cell(row=1, column=1, value="How to Calculate Distance Between Two GPS Coordinates in Excel")
c.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
c.fill = PatternFill("solid", start_color="2E4057")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# ── WHY 30m SECTION ───────────────────────────────────────────────────────
ws.merge_cells("A3:G3")
c = ws.cell(row=3, column=1, value="WHY 30 METRES?  —  Threshold Reasoning")
c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
c.fill = PatternFill("solid", start_color="1F3864")
c.alignment = Alignment(horizontal="left", vertical="center")
c.border = bdr()
ws.row_dimensions[3].height = 20

why_data = [
    ("Consumer GPS accuracy (phone/tablet)",   "±3–15 m",  "Device hardware limitation"),
    ("Restaurant manually places map pin",      "±5–20 m",  "Human click precision varies"),
    ("Different reference point per platform", "±5–25 m",  "Front door vs. parking vs. centroid"),
    ("COMBINED WORST CASE",                    "≈ 30 m",   "Maximum expected drift for SAME branch"),
    ("Two different restaurants in UAE mall",  "50–100 m", "Minimum gap — must stay BELOW this"),
]
headers4 = ["Source of GPS Drift", "Typical Error", "Reason"]
for col, h in enumerate(headers4, 1):
    s(ws, 4, col, h, bold=True, bg="2E86C1", fg="FFFFFF", align="center")
ws.merge_cells("D4:G4")
s(ws, 4, 4, "", bg="2E86C1")
ws.row_dimensions[4].height = 16

for i, (src, err, reason) in enumerate(why_data, 5):
    is_total = i == 8
    is_gap   = i == 9
    bg = "FFF3CD" if is_total else ("FADBD8" if is_gap else ("F5F5F5" if i % 2 == 0 else "FFFFFF"))
    bold = is_total or is_gap
    s(ws, i, 1, src,    bold=bold, bg=bg)
    s(ws, i, 2, err,    bold=bold, bg=bg, align="center")
    ws.merge_cells(f"C{i}:G{i}")
    s(ws, i, 3, reason, bold=bold, bg=bg, italic=not bold)
    ws.row_dimensions[i].height = 16

# ── FORMULA SECTION ───────────────────────────────────────────────────────
ws.merge_cells("A11:G11")
c = ws.cell(row=11, column=1, value="THE DISTANCE FORMULA  —  Two approaches")
c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
c.fill = PatternFill("solid", start_color="1F3864")
c.alignment = Alignment(horizontal="left", vertical="center")
c.border = bdr()
ws.row_dimensions[11].height = 20

# Approach 1 header
ws.merge_cells("A12:G12")
c = ws.cell(row=12, column=1, value="APPROACH 1  —  Flat-Earth Approximation (fast, accurate to ±0.1% for distances < 10 km — good enough for restaurant matching)")
c.font = Font(name="Arial", bold=True, color="1A5276", size=10)
c.fill = PatternFill("solid", start_color="D6EAF8")
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
c.border = bdr()
ws.row_dimensions[12].height = 20

formula_flat = '=SQRT(((lat2-lat1)*111000)^2 + ((lon2-lon1)*COS(RADIANS((lat1+lat2)/2))*111000)^2)'
ws.merge_cells("A13:G13")
c = ws.cell(row=13, column=1, value=formula_flat)
c.font = Font(name="Courier New", bold=True, color="000080", size=10)
c.fill = PatternFill("solid", start_color="EAF2FF")
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
c.border = bdr()
ws.row_dimensions[13].height = 18

# Explain the parts
parts = [
    ("lat2 - lat1",                  "Difference in latitude (degrees)"),
    ("× 111,000",                    "Converts degrees → metres (1° lat = 111 km everywhere)"),
    ("lon2 - lon1",                  "Difference in longitude (degrees)"),
    ("× COS(RADIANS(avg_lat))",      "Shrinks longitude by cosine — longitude degrees get shorter near poles (UAE ≈ 25° → factor ≈ 0.906)"),
    ("× 111,000",                    "Converts to metres"),
    ("SQRT( north² + east² )",       "Pythagoras — combines north/south + east/west components into straight-line distance"),
]
ws.row_dimensions[14].height = 16
s(ws, 14, 1, "Formula Part", bold=True, bg="2E86C1", fg="FFFFFF")
ws.merge_cells("B14:G14")
s(ws, 14, 2, "What it does", bold=True, bg="2E86C1", fg="FFFFFF")

for i, (part, desc) in enumerate(parts, 15):
    bg = "F5F5F5" if i % 2 == 0 else "FFFFFF"
    s(ws, i, 1, part, bg=bg, fg="000080")
    ws.merge_cells(f"B{i}:G{i}")
    s(ws, i, 2, desc, bg=bg, italic=True)
    ws.row_dimensions[i].height = 15

# Approach 2 header
ws.merge_cells("A22:G22")
c = ws.cell(row=22, column=1, value="APPROACH 2  —  Full Haversine (exact, works globally — use when lat/lon span > 100 km)")
c.font = Font(name="Arial", bold=True, color="1A5276", size=10)
c.fill = PatternFill("solid", start_color="D6EAF8")
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
c.border = bdr()
ws.row_dimensions[22].height = 20

formula_hav = '=2*6371000*ASIN(SQRT(SIN((RADIANS(lat2)-RADIANS(lat1))/2)^2 + COS(RADIANS(lat1))*COS(RADIANS(lat2))*SIN((RADIANS(lon2)-RADIANS(lon1))/2)^2))'
ws.merge_cells("A23:G23")
c = ws.cell(row=23, column=1, value=formula_hav)
c.font = Font(name="Courier New", bold=True, color="000080", size=9)
c.fill = PatternFill("solid", start_color="EAF2FF")
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
c.border = bdr()
ws.row_dimensions[23].height = 18

ws.merge_cells("A24:G24")
c = ws.cell(row=24, column=1, value="6,371,000 = Earth radius in metres  |  ASIN/SIN/COS = spherical trigonometry to handle Earth's curvature")
c.font = Font(name="Arial", italic=True, color="555555", size=9)
c.fill = PatternFill("solid", start_color="F5F5F5")
c.alignment = Alignment(horizontal="left", vertical="center")
c.border = bdr()
ws.row_dimensions[24].height = 15

# ── LIVE DEMO TABLE ───────────────────────────────────────────────────────
ws.merge_cells("A26:G26")
c = ws.cell(row=26, column=1, value="LIVE DEMO  —  Try it yourself (edit the yellow cells, distance recalculates automatically)")
c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
c.fill = PatternFill("solid", start_color="1F3864")
c.alignment = Alignment(horizontal="left", vertical="center")
c.border = bdr()
ws.row_dimensions[26].height = 20

demo_headers = ["", "Point A (Talabat)", "", "Point B (Zomato)", "", "Distance (m)", "Match? (≤30m)"]
for col, h in enumerate(demo_headers, 1):
    s(ws, 27, col, h, bold=True, bg="2E86C1", fg="FFFFFF", align="center")
ws.row_dimensions[27].height = 16

s(ws, 28, 1, "Label", bold=True, bg="F0F0F0", align="center")
s(ws, 28, 2, "Latitude A",  bold=True, bg="F0F0F0", align="center")
s(ws, 28, 3, "Longitude A", bold=True, bg="F0F0F0", align="center")
s(ws, 28, 4, "Latitude B",  bold=True, bg="F0F0F0", align="center")
s(ws, 28, 5, "Longitude B", bold=True, bg="F0F0F0", align="center")
s(ws, 28, 6, "Distance (m)", bold=True, bg="F0F0F0", align="center")
s(ws, 28, 7, "Match?",      bold=True, bg="F0F0F0", align="center")
ws.row_dimensions[28].height = 16

demo_rows = [
    ("Al Banosh Seafood",          25.2174724,  55.3153729,  25.2173665303, 55.3153829277),
    ("Emirates Sea Restaurant",    25.2319419,  55.3868504,  25.2320413521, 55.3868798912),
    ("Prepmeal",                   25.2335136,  55.2806796,  25.2334716453, 55.2806597203),
    ("Zaatar W Zeit (chain!)",     25.2086335,  55.2724549,  25.1972087027, 55.2771074697),
    ("B60 Burgers (chain!)",       25.269637,   55.317528,   25.2395999398, 55.3102377802),
    ("Your test  →",              25.2000000,  55.2000000,  25.2000000,    55.2000000),
]

INPUT_BG = "FFF9C4"  # yellow for editable cells

for i, (label, la, loa, lb, lob) in enumerate(demo_rows, 29):
    is_test = i == 34
    bg_in = INPUT_BG if is_test else "FFFFFF"
    s(ws, i, 1, label, bg="F5F5F5", align="left")

    for col, val in [(2, la), (3, loa), (4, lb), (5, lob)]:
        c = ws.cell(row=i, column=col, value=val)
        c.font = Font(name="Arial", size=10, color="000080" if not is_test else "000000")
        c.fill = PatternFill("solid", start_color=bg_in)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = bdr()
        c.number_format = "0.0000000"

    # Flat-earth distance formula referencing this row's cells
    dist_formula = f"=SQRT(((E{i}-C{i})*111000)^2+((D{i}-B{i})*COS(RADIANS((C{i}+E{i})/2))*111000)^2)"
    # Wait, columns: B=lat_a, C=lon_a, D=lat_b, E=lon_b → F=dist, G=match
    # Actually let me re-check: col2=B=lat_a, col3=C=lon_a, col4=D=lat_b, col5=E=lon_b
    col_la = get_column_letter(2)  # B
    col_loa = get_column_letter(3) # C
    col_lb = get_column_letter(4)  # D
    col_lob = get_column_letter(5) # E
    dist_f = f"=SQRT((({col_lb}{i}-{col_la}{i})*111000)^2+(({col_lob}{i}-{col_loa}{i})*COS(RADIANS(({col_la}{i}+{col_lb}{i})/2))*111000)^2)"
    match_f = f'=IF(F{i}<=30,"MATCH ✓","NO MATCH")'

    dc = ws.cell(row=i, column=6, value=dist_f)
    dc.font = Font(name="Arial", size=10, bold=True, color="000000")
    dc.fill = PatternFill("solid", start_color="EAF2FF")
    dc.alignment = Alignment(horizontal="right", vertical="center")
    dc.border = bdr()
    dc.number_format = "#,##0.0"

    mc = ws.cell(row=i, column=7, value=match_f)
    mc.font = Font(name="Arial", size=10, bold=True, color="000000")
    mc.fill = PatternFill("solid", start_color="F5F5F5")
    mc.alignment = Alignment(horizontal="center", vertical="center")
    mc.border = bdr()

    ws.row_dimensions[i].height = 16

# Note about test row
ws.merge_cells("A35:G35")
c = ws.cell(row=35, column=1, value="^ Edit the yellow row (row 34) with any two coordinates to test your own pairs. Distance and Match columns update automatically.")
c.font = Font(name="Arial", italic=True, color="555555", size=9)
c.fill = PatternFill("solid", start_color="FFF9C4")
c.alignment = Alignment(horizontal="left", vertical="center")
c.border = bdr()
ws.row_dimensions[35].height = 15

wb.save(PATH)
print("Sheet added and saved.")
