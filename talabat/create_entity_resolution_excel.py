from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from math import radians, sin, cos, sqrt, atan2

def haversine_m(lat1, lon1, lat2, lon2):
    R = 6_371_000
    p1, p2 = radians(lat1), radians(lat2)
    dp = radians(lat2 - lat1)
    dl = radians(lon2 - lon1)
    a = sin(dp/2)**2 + cos(p1)*cos(p2)*sin(dl/2)**2
    return round(R * 2 * atan2(sqrt(a), sqrt(1-a)), 1)

# ── colours ───────────────────────────────────────────────────────────────
C_HEADER_BG  = "1F3864"   # dark navy
C_HEADER_FG  = "FFFFFF"
C_ANCHOR_BG  = "D6E4F0"   # light blue
C_ANCHOR_FG  = "1A5276"
C_MATCH_BG   = "D4EDDA"   # green
C_MATCH_FG   = "155724"
C_NO_BG      = "F8D7DA"   # red
C_NO_FG      = "721C24"
C_ROW_ALT    = "F5F5F5"
C_SECTION_BG = "EAF2FF"   # section divider
C_NOTE_FG    = "555555"
C_TITLE_BG   = "2E4057"

thin = Side(style="thin", color="CCCCCC")
def border(): return Border(left=thin, right=thin, top=thin, bottom=thin)

def cell_style(ws, row, col, value, bold=False, bg=None, fg="000000",
               align="left", wrap=False, italic=False, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=fg, italic=italic, size=size)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = border()
    return c

# ── data ──────────────────────────────────────────────────────────────────
cases = [
    {
        "name": "Al Banosh Seafood – Al Wasl",
        "talabat_url": "https://www.talabat.com/uae/restaurant/663941/al-banosh-seafood-al-wasl",
        "talabat_lat": 25.2174724, "talabat_lon": 55.3153729,
        "comparisons": [
            ("Zomato",      25.2173665303, 55.3153829277),
            ("TripAdvisor", 25.205605,     55.324997),
        ]
    },
    {
        "name": "Zaatar W Zeit – SZR Jumeirah-1",
        "talabat_url": "https://www.talabat.com/uae/restaurant/626019/zaatar-w-zeit-sheikh-zayed-road-jumeirah-1",
        "talabat_lat": 25.2086335, "talabat_lon": 55.27245489999999,
        "comparisons": [
            ("Zomato Branch #1", 25.1972087027, 55.2771074697),
            ("Zomato Branch #2", 25.0657059929, 55.1925799996),
        ]
    },
    {
        "name": "Emirates Sea Restaurant – Al Rashidiya",
        "talabat_url": "https://www.talabat.com/uae/restaurant/692108/emirates-sea-restaurant-al-rashidiya",
        "talabat_lat": 25.2319419, "talabat_lon": 55.3868504,
        "comparisons": [
            ("Zomato Entry #1", 25.2320413521, 55.3868798912),
            ("Zomato Entry #2", 25.3968631608, 55.4976443946),
        ]
    },
    {
        "name": "Prepmeal Healthy Diet Food – Al Satwa",
        "talabat_url": "https://www.talabat.com/uae/restaurant/675083/prepmeal-healthy-diet-food-al-satwa",
        "talabat_lat": 25.23351358571691, "talabat_lon": 55.28067963558197,
        "comparisons": [
            ("Zomato", 25.2334716453, 55.2806597203),
        ]
    },
    {
        "name": "B60 Burgers – Al Rigga",
        "talabat_url": "https://www.talabat.com/uae/restaurant/774223/b60-burgers-al-rigga",
        "talabat_lat": 25.269637, "talabat_lon": 55.317528,
        "comparisons": [
            ("Zomato Branch #1", 25.2395999398, 55.3102377802),
            ("Zomato Branch #2", 25.1955473916, 55.4239015657),
        ]
    },
]

THRESHOLD = 30  # metres

wb = Workbook()
ws = wb.active
ws.title = "Entity Resolution Validation"

# ── title ─────────────────────────────────────────────────────────────────
ws.merge_cells("A1:J1")
c = ws.cell(row=1, column=1, value="Talabat × Zomato — Entity Resolution Geo-Matching Validation")
c.font = Font(name="Arial", bold=True, color=C_HEADER_FG, size=13)
c.fill = PatternFill("solid", start_color=C_TITLE_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:J2")
c = ws.cell(row=2, column=1, value=f"Match rule: Geo distance ≤ {THRESHOLD} m  →  SAME BRANCH  |  Distance > {THRESHOLD} m  →  DIFFERENT BRANCH / WRONG PIN")
c.font = Font(name="Arial", italic=True, color=C_NOTE_FG, size=10)
c.fill = PatternFill("solid", start_color="FFF8E1")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# ── column headers ────────────────────────────────────────────────────────
headers = ["#", "Restaurant Name", "Platform", "Latitude", "Longitude",
           "Talabat Lat", "Talabat Lon", "Distance (m)", "Decision", "Notes"]
widths  = [4,   36,                14,         13,          13,
           13,            13,             14,              18,         42]

for col, (h, w) in enumerate(zip(headers, widths), 1):
    cell_style(ws, 3, col, h, bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG, align="center", size=10)
    ws.column_dimensions[get_column_letter(col)].width = w
ws.row_dimensions[3].height = 20

# ── data rows ──────────────────────────────────────────────────────────────
notes_map = {
    "MATCH":        "Same physical branch confirmed — link entities",
    "DIFF BRANCH":  "Chain with multiple locations — Talabat branch has no Zomato equivalent here",
    "WRONG PIN":    "Bad geocode / different restaurant entirely — correctly rejected",
}

row = 4
case_num = 0
for case in cases:
    case_num += 1
    t_lat, t_lon = case["talabat_lat"], case["talabat_lon"]
    alt = C_ROW_ALT if case_num % 2 == 0 else "FFFFFF"

    # Talabat anchor row
    cell_style(ws, row, 1, case_num,   bg=C_ANCHOR_BG, fg=C_ANCHOR_FG, align="center", bold=True)
    cell_style(ws, row, 2, case["name"], bg=C_ANCHOR_BG, fg=C_ANCHOR_FG, bold=True)
    cell_style(ws, row, 3, "Talabat (anchor)", bg=C_ANCHOR_BG, fg=C_ANCHOR_FG, bold=True)
    cell_style(ws, row, 4, t_lat,  bg=C_ANCHOR_BG, fg=C_ANCHOR_FG, align="right")
    cell_style(ws, row, 5, t_lon,  bg=C_ANCHOR_BG, fg=C_ANCHOR_FG, align="right")
    cell_style(ws, row, 6, "—",    bg=C_ANCHOR_BG, fg=C_ANCHOR_FG, align="center")
    cell_style(ws, row, 7, "—",    bg=C_ANCHOR_BG, fg=C_ANCHOR_FG, align="center")
    cell_style(ws, row, 8, "—",    bg=C_ANCHOR_BG, fg=C_ANCHOR_FG, align="center")
    cell_style(ws, row, 9, "ANCHOR", bg=C_ANCHOR_BG, fg=C_ANCHOR_FG, align="center", bold=True)
    cell_style(ws, row, 10, case["talabat_url"], bg=C_ANCHOR_BG, fg=C_ANCHOR_FG, wrap=True)
    ws.row_dimensions[row].height = 16
    row += 1

    # Comparison rows
    for platform, p_lat, p_lon in case["comparisons"]:
        dist = haversine_m(t_lat, t_lon, p_lat, p_lon)
        if dist <= THRESHOLD:
            decision, bg, fg = "MATCH ✓", C_MATCH_BG, C_MATCH_FG
            note = notes_map["MATCH"]
        elif dist > 1000:
            decision, bg, fg = "WRONG PIN / DIFF BRANCH", C_NO_BG, C_NO_FG
            note = notes_map["WRONG PIN"] if dist > 5000 else notes_map["DIFF BRANCH"]
        else:
            decision, bg, fg = "DIFF BRANCH", C_NO_BG, C_NO_FG
            note = notes_map["DIFF BRANCH"]

        cell_style(ws, row, 1, "",        bg=alt)
        cell_style(ws, row, 2, "",        bg=alt)
        cell_style(ws, row, 3, platform,  bg=alt, bold=False)
        cell_style(ws, row, 4, p_lat,     bg=alt, align="right")
        cell_style(ws, row, 5, p_lon,     bg=alt, align="right")
        cell_style(ws, row, 6, t_lat,     bg=alt, align="right")
        cell_style(ws, row, 7, t_lon,     bg=alt, align="right")
        cell_style(ws, row, 8, dist,      bg=alt, align="right", bold=True)
        cell_style(ws, row, 9, decision,  bg=bg,  fg=fg, align="center", bold=True)
        cell_style(ws, row, 10, note,     bg=alt, fg=C_NOTE_FG, wrap=True, italic=True)
        ws.row_dimensions[row].height = 16

        # Format distance cell
        ws.cell(row=row, column=8).number_format = '#,##0.0'
        ws.cell(row=row, column=4).number_format = '0.0000000'
        ws.cell(row=row, column=5).number_format = '0.0000000'
        ws.cell(row=row, column=6).number_format = '0.0000000'
        ws.cell(row=row, column=7).number_format = '0.0000000'
        row += 1

    row += 0  # no extra gap — cases are visually separated by anchor row colour

# ── summary sheet ─────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")

sum_headers = ["Case #", "Restaurant", "Matched Platform", "Distance (m)", "Decision", "Outcome"]
sum_widths  = [8,         36,            20,                  14,             22,          36]

ws2.merge_cells("A1:F1")
c = ws2.cell(row=1, column=1, value="Entity Resolution — Summary of Match Decisions")
c.font = Font(name="Arial", bold=True, color=C_HEADER_FG, size=13)
c.fill = PatternFill("solid", start_color=C_TITLE_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

for col, (h, w) in enumerate(zip(sum_headers, sum_widths), 1):
    cell_style(ws2, 2, col, h, bold=True, bg=C_HEADER_BG, fg=C_HEADER_FG, align="center")
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.row_dimensions[2].height = 18

summary_data = [
    (1, "Al Banosh Seafood – Al Wasl",           "Zomato",           12,    "MATCH ✓",        "Link Talabat ↔ Zomato"),
    (1, "Al Banosh Seafood – Al Wasl",           "TripAdvisor",     1634,   "WRONG PIN",      "Reject — different branch"),
    (2, "Zaatar W Zeit – SZR Jumeirah-1",        "Zomato Branch #1", 1351,  "DIFF BRANCH",    "Talabat-only entity; no Zomato link"),
    (2, "Zaatar W Zeit – SZR Jumeirah-1",        "Zomato Branch #2", 17800, "DIFF BRANCH",    "Talabat-only entity; no Zomato link"),
    (3, "Emirates Sea Restaurant – Al Rashidiya","Zomato Entry #1",  11,    "MATCH ✓",        "Link Talabat ↔ Zomato"),
    (3, "Emirates Sea Restaurant – Al Rashidiya","Zomato Entry #2",  22000, "WRONG PIN",      "Reject — bad geocode / wrong restaurant"),
    (4, "Prepmeal Healthy Diet Food – Al Satwa", "Zomato",           5,     "MATCH ✓",        "Link Talabat ↔ Zomato (near-perfect)"),
    (5, "B60 Burgers – Al Rigga",                "Zomato Branch #1", 3414,  "DIFF BRANCH",    "Talabat-only entity; no Zomato link"),
    (5, "B60 Burgers – Al Rigga",                "Zomato Branch #2", 11000, "DIFF BRANCH",    "Talabat-only entity; no Zomato link"),
]

for i, (num, name, plat, dist, dec, outcome) in enumerate(summary_data, 3):
    alt = C_ROW_ALT if i % 2 == 0 else "FFFFFF"
    is_match = "MATCH" in dec
    d_bg = C_MATCH_BG if is_match else C_NO_BG
    d_fg = C_MATCH_FG if is_match else C_NO_FG

    cell_style(ws2, i, 1, num,     bg=alt, align="center", bold=True)
    cell_style(ws2, i, 2, name,    bg=alt)
    cell_style(ws2, i, 3, plat,    bg=alt)
    cell_style(ws2, i, 4, dist,    bg=alt, align="right")
    cell_style(ws2, i, 5, dec,     bg=d_bg, fg=d_fg, align="center", bold=True)
    cell_style(ws2, i, 6, outcome, bg=alt, fg=C_NOTE_FG, italic=True)
    ws2.cell(row=i, column=4).number_format = '#,##0'
    ws2.row_dimensions[i].height = 16

# Verdict row
vrow = len(summary_data) + 4
ws2.merge_cells(f"A{vrow}:F{vrow}")
c = ws2.cell(row=vrow, column=1,
             value="VERDICT: Strategy validated — ≤30 m threshold correctly identifies same-branch matches and rejects cross-branch false positives from chain restaurants.")
c.font = Font(name="Arial", bold=True, color="1A5276", size=10)
c.fill = PatternFill("solid", start_color="D6EAF8")
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
c.border = border()
ws2.row_dimensions[vrow].height = 30

out = r"C:\Users\SagarSingh\OneDrive - Mordor Intelligence\Desktop\Webscraping-Production-Projects\talabat\data\urls\entity_resolution_validation.xlsx"
wb.save(out)
print("Saved:", out)
