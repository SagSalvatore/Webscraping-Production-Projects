"""
Talabat UAE — Menu Tracker with Delta Engine
=============================================
Scrapes current menu data from Talabat restaurant pages and compares against
the baseline (Data_menus.xlsx, ~1 year old data) using hashlib + difflib.

Pipeline:
  1. build_baseline()  — group Excel rows by branch_id → baseline JSON
  2. scrape_menu()     — curl_cffi + Oxylabs proxy → __NEXT_DATA__ menu
  3. delta_engine()    — hashlib (fast gate) + difflib (deep diff per item)
  4. write_outputs()   — JSON snapshot + Excel delta report (5 sheets)

Run modes:
  TEST_MODE = True   → first 200 unique restaurants only
  TEST_MODE = False  → all 5,275 restaurants

Usage:
  python talabat_menu_tracker.py
  python talabat_menu_tracker.py --all      # full run
  python talabat_menu_tracker.py --rebuild  # force rebuild baseline
"""

import sys
# Force UTF-8 stdout on Windows so emoji/arrows in print() don't crash
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import argparse
import csv
import difflib
import hashlib
import json
import os
import re
import random
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from bs4 import BeautifulSoup, NavigableString
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from curl_cffi import requests as cffi_requests

# ── paths ─────────────────────────────────────────────────────────────────────
ROOT       = Path(__file__).resolve().parent.parent          # talabat/
ENV_PATH   = ROOT / ".env"
EXCEL_SRC  = ROOT / "Data_menus.xlsx"
DATA_DIR   = Path(__file__).resolve().parent / "data"
BASELINE_F = DATA_DIR / "baseline" / "baseline.json"
SNAP_DIR   = DATA_DIR / "snapshots"
REPORT_DIR = DATA_DIR / "reports"

# ── settings ──────────────────────────────────────────────────────────────────
TEST_LIMIT           = 200    # restaurants in test mode
CONCURRENT_WORKERS   = 8
MIN_DELAY            = 1.5
MAX_DELAY            = 3.0
RETRY_ATTEMPTS       = 3
SESSION_REFRESH      = 30
SAVE_INTERVAL        = 20     # checkpoint every N restaurants

# ── credentials ───────────────────────────────────────────────────────────────
load_dotenv(dotenv_path=ENV_PATH)
USERNAME = os.getenv("OXYLABS_USERNAME", "")
PASSWORD = os.getenv("OXYLABS_PASSWORD", "")
COUNTRY  = os.getenv("OXYLABS_COUNTRY", "ae")

# ── thread-safe accumulators ──────────────────────────────────────────────────
lock          = threading.Lock()
all_snapshots = []          # current scrape results
all_deltas    = []          # delta records
processed_ids = set()
counters      = {"scraped": 0, "changed": 0, "no_change": 0, "failed": 0, "new": 0}

BROWSER_UAS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — BASELINE BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

def parse_price_aed(price_str) -> float | None:
    """Convert 'AED 48', 'AED 48.5', NaN → float or None."""
    if not price_str or pd.isna(price_str):
        return None
    m = re.search(r"[\d]+\.?[\d]*", str(price_str))
    return float(m.group()) if m else None


def make_item_key(name: str) -> str:
    """Normalize item name for matching: lowercase, collapse spaces, strip."""
    return re.sub(r"\s+", " ", str(name).lower().strip())


def compute_hash(data: dict | list) -> str:
    """SHA-256 of deterministically serialized data."""
    payload = json.dumps(data, sort_keys=True, ensure_ascii=False)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def build_baseline(force: bool = False) -> dict:
    """
    Build baseline from Data_menus.xlsx grouped by branch_id.
    Returns dict keyed by branch_id (int).
    Cached to baseline/baseline.json — rebuild with force=True.
    """
    if BASELINE_F.exists() and not force:
        print(f"Loading cached baseline from {BASELINE_F}")
        with open(BASELINE_F, "r", encoding="utf-8") as f:
            raw = json.load(f)
        return {int(k): v for k, v in raw.items()}

    print(f"Building baseline from {EXCEL_SRC} ...")
    df = pd.read_excel(EXCEL_SRC, dtype={"branch_id": int, "restaurant_id": int})
    df["price_aed"] = df["Price"].apply(parse_price_aed)
    df["item_key"]  = df["Menu item(name)"].apply(
        lambda x: make_item_key(x) if pd.notna(x) else ""
    )

    baseline = {}
    for branch_id, grp in df.groupby("branch_id"):
        items = []
        for _, row in grp.iterrows():
            if not row["item_key"]:
                continue
            items.append({
                "item_name":   str(row["Menu item(name)"]).strip() if pd.notna(row["Menu item(name)"]) else "",
                "item_key":    row["item_key"],
                "category":    str(row["Menu category"]).strip()    if pd.notna(row["Menu category"])    else "",
                "price_aed":   row["price_aed"],
                "description": str(row["DESCRIPTION"]).strip()      if pd.notna(row["DESCRIPTION"])      else "",
            })

        # price hash: {item_key: price} — detects any AED change fast
        # full hash : price + category + description — catches ALL 4 field types
        price_map = {it["item_key"]: it["price_aed"] for it in items if it["price_aed"] is not None}
        full_map  = [
            {"k": it["item_key"], "p": it["price_aed"],
             "c": it["category"], "d": it.get("description", "")}
            for it in items
        ]

        row0 = grp.iloc[0]
        baseline[int(branch_id)] = {
            "branch_id":       int(branch_id),
            "restaurant_id":   int(row0["restaurant_id"]) if pd.notna(row0["restaurant_id"]) else None,
            "restaurant_name": str(row0["RESTRO NAME"]).strip(),
            "url":             str(row0["map_url"]).strip(),
            "area_name":       str(row0["area_name"]).strip() if pd.notna(row0["area_name"]) else "",
            "serves_cuisine":  str(row0["serves_cuisine"]).strip() if pd.notna(row0["serves_cuisine"]) else "",
            "baseline_items":  items,
            "baseline_price_hash": compute_hash(price_map),
            "baseline_full_hash":  compute_hash(full_map),
            "baseline_item_count": len(items),
            "baseline_scraped_at": None,  # unknown — data is ~1 year old
        }

    BASELINE_F.parent.mkdir(parents=True, exist_ok=True)
    with open(BASELINE_F, "w", encoding="utf-8") as f:
        json.dump(baseline, f, ensure_ascii=False)
    print(f"Baseline built: {len(baseline)} restaurants -> {BASELINE_F}")
    return baseline


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — SCRAPER
# ═══════════════════════════════════════════════════════════════════════════════

def create_session():
    if not USERNAME or not PASSWORD:
        raise RuntimeError("Oxylabs credentials missing. Check talabat/.env")
    sid = random.randint(10000, 99999)
    proxy_url = f"http://customer-{USERNAME}-cc-{COUNTRY}-sessid-{sid}:{PASSWORD}@pr.oxylabs.io:7777"
    session = cffi_requests.Session(impersonate="chrome124")
    session.proxies = {"http": proxy_url, "https": proxy_url}
    session.headers.update({
        "User-Agent":      random.choice(BROWSER_UAS),
        "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Cache-Control":   "no-cache",
        "Pragma":          "no-cache",
    })
    return session, sid


def extract_page_name(soup: "BeautifulSoup") -> str:
    """
    Extract the live restaurant name from the page.
    Talabat HTML structure (verified via DevTools):
      <a data-testid="branch-name">
        <h1 data-testid="restaurant-title">
          "The Monk"   ← direct NavigableString (what we want)
          <br>
          <small class="light-text">in</small>
          ...area...
        </h1>
      </a>
    We extract only the direct text nodes of h1, ignoring <small> children.
    """
    h1 = soup.select_one('h1[data-testid="restaurant-title"]')
    if h1:
        # Direct text nodes only — excludes <br> and <small>in...</small>
        text_nodes = [str(c).strip() for c in h1.children
                      if isinstance(c, NavigableString) and str(c).strip()]
        name = " ".join(text_nodes).strip()
        if name:
            return name
    # Fallback: og:title — strip Talabat's "inArea,UAE" suffix
    og = soup.find("meta", property="og:title")
    if og and og.get("content"):
        raw = str(og["content"]).strip()
        cleaned = re.sub(r'in[A-Z][^,]*,.{2,6}$', '', raw).strip()
        return cleaned if cleaned else raw
    return ""


def parse_next_data_menu(soup: "BeautifulSoup") -> list[dict]:
    """
    Extract menu items from __NEXT_DATA__ SSR JSON embedded in the page.
    Uses BeautifulSoup tag lookup (more robust than regex for malformed HTML).
    Talabat web SSR stores price in AED directly (not fils).
    Returns list of dicts: item_id, item_name, item_key, category,
    price_aed, description, image_url.
    """
    script_tag = soup.find("script", {"id": "__NEXT_DATA__"})
    if not script_tag or not script_tag.string:
        return []

    try:
        nd = json.loads(script_tag.string)
    except (json.JSONDecodeError, ValueError):
        return []

    try:
        page_props = nd["props"]["pageProps"]
        menu_state = page_props["initialMenuState"]["menuData"]
        raw_items  = menu_state.get("items", [])
    except (KeyError, TypeError):
        return []

    items = []
    for it in raw_items:
        if not isinstance(it, dict):
            continue
        price_aed = round(float(it.get("price", 0) or 0), 2)
        item_name = str(it.get("name", "")).strip()
        items.append({
            "item_id":     it.get("id"),
            "item_name":   item_name,
            "item_key":    make_item_key(item_name),
            "category":    str(it.get("originalSection", "")).strip(),
            "price_aed":   price_aed,
            "description": str(it.get("description", "")).strip(),
            "image_url":   it.get("squareImageUrl") or it.get("imageUrl") or "",
        })
    return items


def fetch_menu(session, url: str, worker_id: int) -> dict:
    """
    Fetch one restaurant page and return parsed menu + live name.
    Upgrades from TALABAT_FAST_MENU.py:
      - URL anchor strip (url.split('#')[0])
      - Soft rate-limit text detection ('rate limit' in body)
      - BeautifulSoup parse reused for both name and __NEXT_DATA__
      - raise_for_status equivalent for non-200/429 codes
    """
    try:
        clean_url = url.split("#")[0].strip()
        resp = session.get(clean_url, timeout=25)

        # Hard rate-limit (HTTP 429) or soft rate-limit (text in body)
        if resp.status_code == 429 or "rate limit" in resp.text.lower():
            return {"status": "rate_limited"}

        if resp.status_code != 200:
            return {"status": "http_error", "http_code": resp.status_code}

        # Parse once, reuse for name + menu
        soup  = BeautifulSoup(resp.text, "html.parser")
        name  = extract_page_name(soup)
        items = parse_next_data_menu(soup)
        return {"status": "ok", "items": items, "page_name": name}

    except Exception as exc:
        return {"status": "exception", "error": str(exc)[:200]}


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — DELTA ENGINE (hashlib gate + difflib deep diff)
# ═══════════════════════════════════════════════════════════════════════════════

def run_delta(baseline: dict, new_items: list[dict], branch_id: int) -> dict:
    """
    Compare new scrape against baseline.
    Returns delta dict with added/removed/changed item lists + status.
    """
    base_items = baseline.get("baseline_items", [])
    base_ph    = baseline.get("baseline_price_hash", "")
    base_fh    = baseline.get("baseline_full_hash", "")

    # Layer 1 hashes — must match what build_baseline() computes
    # price_hash: only prices → detects AED changes fast
    # full_hash : price + category + description → catches ALL field changes
    new_price_map = {it["item_key"]: it["price_aed"] for it in new_items if it["price_aed"] is not None}
    new_full_map  = [
        {"k": it["item_key"], "p": it["price_aed"],
         "c": it["category"], "d": it.get("description", "")}
        for it in new_items
    ]
    new_ph = compute_hash(new_price_map)
    new_fh = compute_hash(new_full_map)

    # Fast gate
    if new_ph == base_ph and new_fh == base_fh:
        return {
            "status":      "no_change",
            "new_ph":      new_ph,
            "new_fh":      new_fh,
            "added":       [],
            "removed":     [],
            "changed":     [],
            "price_delta": 0.0,
        }

    # Deep diff — build lookup dicts keyed by item_key
    old_map = {it["item_key"]: it for it in base_items if it["item_key"]}
    new_map = {it["item_key"]: it for it in new_items  if it["item_key"]}

    added   = [new_map[k] for k in new_map if k not in old_map]
    removed = [old_map[k] for k in old_map if k not in new_map]
    changed = []

    for key in old_map:
        if key not in new_map:
            continue
        old = old_map[key]
        new = new_map[key]
        field_changes = {}

        for field in ("price_aed", "category", "description"):
            old_val = old.get(field)
            new_val = new.get(field)
            if field == "price_aed":
                # Float compare with 0.01 AED tolerance
                if old_val is not None and new_val is not None:
                    if abs(float(old_val) - float(new_val)) > 0.01:
                        field_changes[field] = {"old": old_val, "new": new_val}
                elif old_val != new_val:
                    field_changes[field] = {"old": old_val, "new": new_val}
            else:
                if str(old_val or "").strip() != str(new_val or "").strip():
                    field_changes[field] = {"old": old_val, "new": new_val}

        if field_changes:
            changed.append({
                "item_key":     key,
                "item_name":    new.get("item_name", old.get("item_name", "")),
                "category":     new.get("category", ""),
                "field_changes": field_changes,
            })

    # Unified text diff — includes all 4 fields: name(key), price, category, description
    def _row(it):
        return {
            "item": it.get("item_key", ""),
            "price_aed": it.get("price_aed"),
            "category": it.get("category", ""),
            "description": (it.get("description") or "")[:80],  # trim long descriptions
        }
    old_lines = json.dumps(
        sorted([_row(it) for it in base_items], key=lambda x: x["item"]),
        indent=1
    ).splitlines(keepends=True)
    new_lines = json.dumps(
        sorted([_row(it) for it in new_items], key=lambda x: x["item"]),
        indent=1
    ).splitlines(keepends=True)
    text_diff = "".join(difflib.unified_diff(
        old_lines, new_lines,
        fromfile="baseline_~1yr", tofile="current",
        lineterm=""
    ))[:2000]

    # Per-field change counts for the summary report
    price_changes_count       = sum(1 for c in changed if "price_aed"   in c["field_changes"])
    description_changes_count = sum(1 for c in changed if "description" in c["field_changes"])
    category_changes_count    = sum(1 for c in changed if "category"    in c["field_changes"])

    # Total price delta (sum of absolute AED changes)
    price_delta = sum(
        abs((ch["field_changes"].get("price_aed", {}).get("new", 0) or 0) -
            (ch["field_changes"].get("price_aed", {}).get("old", 0) or 0))
        for ch in changed
        if "price_aed" in ch["field_changes"]
    )

    return {
        "status":                   "changed",
        "new_ph":                   new_ph,
        "new_fh":                   new_fh,
        "added":                    added,
        "removed":                  removed,
        "changed":                  changed,
        "text_diff":                text_diff,
        "price_delta":              round(price_delta, 2),
        "price_changes_count":      price_changes_count,
        "description_changes_count": description_changes_count,
        "category_changes_count":   category_changes_count,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — WORKER
# ═══════════════════════════════════════════════════════════════════════════════

CHECKPOINT_F = DATA_DIR / "checkpoint.json"


def load_checkpoint() -> set:
    if CHECKPOINT_F.exists():
        with open(CHECKPOINT_F, "r", encoding="utf-8") as f:
            data = json.load(f)
        return set(data.get("done", []))
    return set()


def save_checkpoint(done: set):
    tmp = str(CHECKPOINT_F) + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump({"done": list(done), "count": len(done),
                   "saved_at": datetime.now(timezone.utc).isoformat()}, f)
    os.replace(tmp, CHECKPOINT_F)


def worker(worker_id: int, jobs: list[tuple], baseline: dict):
    """
    jobs: list of (branch_id, url) tuples.
    baseline: full baseline dict.
    """
    session = None
    local_n = 0

    try:
        for i, (branch_id, url) in enumerate(jobs):
            # Session create / rotate
            if session is None or local_n % SESSION_REFRESH == 0:
                if session:
                    session.close()
                session, sid = create_session()
                verb = "created" if local_n == 0 else "rotated"
                print(f"[W{worker_id}] Session {verb} (sid={sid})")

            print(f"[W{worker_id}] ({i+1}/{len(jobs)}) branch={branch_id}")
            scraped_at = datetime.now(timezone.utc).isoformat()

            result = None
            for attempt in range(RETRY_ATTEMPTS):
                result = fetch_menu(session, url, worker_id)

                if result["status"] == "rate_limited":
                    # 30s base backoff (from TALABAT_FAST_MENU.py) → 30/60/120s
                    wait = (2 ** attempt) * 30
                    print(f"[W{worker_id}] Rate limited — wait {wait}s, rotate session")
                    time.sleep(wait)
                    if session: session.close()
                    session, sid = create_session()
                    continue

                if result["status"] == "exception" and attempt < RETRY_ATTEMPTS - 1:
                    print(f"[W{worker_id}] Retry {attempt+2}/{RETRY_ATTEMPTS}: {result.get('error','')[:60]}")
                    time.sleep(5)
                    continue
                break

            bline = baseline.get(int(branch_id), {})
            rest_name = bline.get("restaurant_name", f"branch_{branch_id}")

            if result["status"] != "ok":
                snap = {
                    "branch_id":      branch_id,
                    "restaurant_name": rest_name,
                    "url":            url,
                    "status":         result["status"],
                    "error":          result.get("error") or result.get("http_code"),
                    "scraped_at":     scraped_at,
                }
                with lock:
                    all_snapshots.append(snap)
                    processed_ids.add(branch_id)
                    counters["failed"] += 1
                    _maybe_checkpoint()
                local_n += 1
                time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))
                continue

            new_items = result["items"]
            # page_name is the LIVE name scraped from the page (may differ from baseline)
            page_name = result.get("page_name", "") or ""

            # Detect name change vs baseline
            name_changed = (
                page_name and rest_name
                and page_name.lower().strip() != rest_name.lower().strip()
            )

            # Run delta only if baseline exists for this branch
            if bline:
                delta = run_delta(bline, new_items, branch_id)
                delta_status = delta["status"]
            else:
                delta = {"status": "new_restaurant", "added": new_items,
                         "removed": [], "changed": [], "price_delta": 0.0,
                         "new_ph": compute_hash({}), "new_fh": compute_hash([])}
                delta_status = "new"

            # Build snapshot
            price_map = {it["item_key"]: it["price_aed"] for it in new_items if it["price_aed"] is not None}
            full_map  = [{"k": it["item_key"], "p": it["price_aed"], "c": it["category"]} for it in new_items]

            snap = {
                "branch_id":        branch_id,
                "restaurant_id":    bline.get("restaurant_id"),
                "restaurant_name":  rest_name,       # from baseline (Excel)
                "page_name":        page_name,        # live from page (FAST_MENU upgrade)
                "name_changed":     name_changed,     # True if name differs from baseline
                "url":              url,
                "area_name":        bline.get("area_name", ""),
                "serves_cuisine":   bline.get("serves_cuisine", ""),
                "scraped_at":       scraped_at,
                "status":           delta_status,
                "item_count":       len(new_items),
                "baseline_count":   bline.get("baseline_item_count", 0),
                "price_hash":       delta.get("new_ph", ""),
                "full_hash":        delta.get("new_fh", ""),
                "items":            new_items,
                "delta": {
                    "status":                    delta_status,
                    "added_count":               len(delta.get("added", [])),
                    "removed_count":             len(delta.get("removed", [])),
                    "changed_count":             len(delta.get("changed", [])),
                    "price_changes_count":       delta.get("price_changes_count", 0),
                    "description_changes_count": delta.get("description_changes_count", 0),
                    "category_changes_count":    delta.get("category_changes_count", 0),
                    "price_delta_aed":           delta.get("price_delta", 0),
                    "added":                     delta.get("added", []),
                    "removed":                   delta.get("removed", []),
                    "changed":                   delta.get("changed", []),
                    "text_diff":                 delta.get("text_diff", ""),
                },
            }

            with lock:
                all_snapshots.append(snap)
                all_deltas.extend(_expand_deltas(snap))
                processed_ids.add(branch_id)
                counters["scraped"] += 1
                counters[delta_status if delta_status in counters else "changed"] += 1
                _maybe_checkpoint()

            name_flag = f" [NAME CHANGED: {rest_name} -> {page_name}]" if name_changed else ""
            print(
                f"[W{worker_id}] {rest_name[:28]} | items={len(new_items)} "
                f"| {delta_status} "
                f"(+{len(delta.get('added',[]))} -{len(delta.get('removed',[]))} ~{len(delta.get('changed',[]))})"
                f"{name_flag}"
            )
            local_n += 1
            time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

    except Exception as exc:
        print(f"[W{worker_id}] Critical: {exc}")
    finally:
        if session:
            session.close()
        print(f"[W{worker_id}] Done. local_n={local_n}")


def _maybe_checkpoint():
    total = sum(counters.values())
    if total % SAVE_INTERVAL == 0:
        save_checkpoint(processed_ids)
        print(f"[checkpoint] total={total} | scraped={counters['scraped']} "
              f"| changed={counters['changed']} | no_change={counters['no_change']} "
              f"| failed={counters['failed']}")


def _expand_deltas(snap: dict) -> list[dict]:
    """Flatten snap delta into individual change rows for Excel output."""
    rows = []
    branch_id   = snap["branch_id"]
    rest_name   = snap["restaurant_name"]
    scraped_at  = snap["scraped_at"]
    delta       = snap["delta"]

    for it in delta.get("added", []):
        rows.append({
            "change_type":      "ADDED",
            "branch_id":        branch_id,
            "restaurant_name":  rest_name,
            "item_name":        it.get("item_name", ""),
            "category":         it.get("category", ""),
            "new_price_aed":    it.get("price_aed"),
            "old_price_aed":    None,
            "price_change_pct": None,
            "field_changed":    "new_item",
            "old_value":        "",
            "new_value":        "",
            "description":      it.get("description", ""),
            "scraped_at":       scraped_at,
        })

    for it in delta.get("removed", []):
        rows.append({
            "change_type":      "REMOVED",
            "branch_id":        branch_id,
            "restaurant_name":  rest_name,
            "item_name":        it.get("item_name", ""),
            "category":         it.get("category", ""),
            "new_price_aed":    None,
            "old_price_aed":    it.get("price_aed"),
            "price_change_pct": None,
            "field_changed":    "removed_item",
            "old_value":        "",
            "new_value":        "",
            "description":      it.get("description", ""),
            "scraped_at":       scraped_at,
        })

    for ch in delta.get("changed", []):
        for field, vals in ch.get("field_changes", {}).items():
            old_v = vals.get("old")
            new_v = vals.get("new")
            pct   = None
            # User-friendly field label
            field_label = {
                "price_aed":   "price",
                "category":    "menu_category",
                "description": "description",
            }.get(field, field)

            if field == "price_aed" and old_v and new_v:
                try:
                    pct = round(((float(new_v) - float(old_v)) / float(old_v)) * 100, 2)
                except ZeroDivisionError:
                    pass
            rows.append({
                "change_type":   "CHANGED",
                "branch_id":     branch_id,
                "restaurant_name": rest_name,
                "item_name":     ch.get("item_name", ch.get("item_key", "")),
                "category":      ch.get("category", ""),
                "new_price_aed": new_v if field == "price_aed" else None,
                "old_price_aed": old_v if field == "price_aed" else None,
                "price_change_pct": pct,
                "field_changed": field_label,   # "price" | "menu_category" | "description"
                "old_value":     str(old_v) if old_v is not None else "",
                "new_value":     str(new_v) if new_v is not None else "",
                "scraped_at":    scraped_at,
            })

    return rows


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — OUTPUT WRITERS
# ═══════════════════════════════════════════════════════════════════════════════

_THIN = Side(style="thin", color="CCCCCC")
_THICK = Side(style="medium", color="AAAAAA")


def _bdr(thick=False):
    s = _THICK if thick else _THIN
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value, bold=False, bg=None, fg="000000",
          align="left", wrap=False, italic=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=fg, italic=italic, size=9)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = _bdr()
    if num_fmt:
        c.number_format = num_fmt
    return c


def write_json(run_id: str):
    """Write full snapshot JSON and compact delta JSON."""
    snap_path  = SNAP_DIR  / f"snapshot_{run_id}.json"
    delta_path = REPORT_DIR / f"delta_{run_id}.json"

    with open(snap_path, "w", encoding="utf-8") as f:
        json.dump(all_snapshots, f, ensure_ascii=False, indent=2)
    print(f"Snapshot JSON: {snap_path}")

    # Compact delta — only changed/added/removed restaurants
    changed_snaps = [s for s in all_snapshots if s.get("status") in ("changed", "new")]
    with open(delta_path, "w", encoding="utf-8") as f:
        json.dump(changed_snaps, f, ensure_ascii=False, indent=2)
    print(f"Delta JSON   : {delta_path}")


def write_excel(run_id: str):
    """
    5-sheet Excel workbook:
      1. Run Summary     — one row per restaurant
      2. Price Changes   — rows where price_aed changed
      3. New Items       — items added vs baseline
      4. Removed Items   — items missing vs baseline
      5. Full Menu       — all current menu items scraped this run
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M UTC")

    # ── colour palette ───────────────────────────────────────────────────────
    HDR_BG = "1F3864"   # dark navy header
    CHANGED_BG = "FFF3CD"   # amber
    ADDED_BG   = "D4EDDA"   # green
    REMOVED_BG = "F8D7DA"   # red
    NOCHANGE_BG= "E8F4FD"   # light blue
    FAIL_BG    = "FADBD8"   # salmon

    # ── Sheet 1: Run Summary ─────────────────────────────────────────────────
    ws1 = wb.create_sheet("Run Summary")
    ws1.freeze_panes = "A3"
    ws1.merge_cells("A1:Q1")
    c = ws1.cell(row=1, column=1,
                 value=f"Talabat Menu Tracker — Run {run_id} | {now_str} | {len(all_snapshots)} restaurants")
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.fill = PatternFill("solid", start_color="2E4057")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 24

    sum_headers = [
        "branch_id", "restaurant_name", "page_name", "name_changed", "area_name", "url", "status",
        "menu_items_before", "menu_items_after",
        "items_added", "items_removed",
        "items_price_changed", "items_desc_changed", "items_cat_changed", "items_any_changed",
        "price_delta_aed", "scraped_at",
    ]
    sum_widths = [12, 32, 32, 13, 18, 52, 14, 16, 16, 13, 14, 18, 17, 16, 16, 15, 22]
    for col, (h, w) in enumerate(zip(sum_headers, sum_widths), 1):
        _cell(ws1, 2, col, h, bold=True, bg=HDR_BG, fg="FFFFFF", align="center")
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[2].height = 18

    status_bg = {"changed": CHANGED_BG, "no_change": NOCHANGE_BG,
                 "new": ADDED_BG, "new_restaurant": ADDED_BG,
                 "failed": FAIL_BG, "http_error": FAIL_BG,
                 "rate_limited": FAIL_BG, "exception": FAIL_BG}

    for i, snap in enumerate(all_snapshots, 3):
        st = snap.get("status", "unknown")
        bg = status_bg.get(st, "FFFFFF")
        alt = "F9F9F9" if i % 2 == 0 else "FFFFFF"
        d  = snap.get("delta", {})
        nc = snap.get("name_changed", False)
        vals = [
            snap.get("branch_id"),
            snap.get("restaurant_name", ""),
            snap.get("page_name", ""),
            "YES" if nc else "",
            snap.get("area_name", ""),
            snap.get("url", ""),
            st.upper(),
            snap.get("baseline_count", 0),   # menu_items_before
            snap.get("item_count", 0),        # menu_items_after
            d.get("added_count", 0),          # items_added
            d.get("removed_count", 0),        # items_removed
            d.get("price_changes_count", 0),  # items_price_changed
            d.get("description_changes_count", 0),  # items_desc_changed
            d.get("category_changes_count", 0),     # items_cat_changed
            d.get("changed_count", 0),        # items_any_changed
            d.get("price_delta_aed", 0),
            snap.get("scraped_at", ""),
        ]
        center_cols = {1, 4, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16}
        for col, val in enumerate(vals, 1):
            name_chg_bg = "FFE0B2" if (col == 4 and nc) else (bg if st in status_bg else alt)
            _cell(ws1, i, col, val,
                  bg=name_chg_bg,
                  align="center" if col in center_cols else "left",
                  num_fmt="#,##0.00" if col == 16 else None)
        ws1.row_dimensions[i].height = 15

    # ── Sheet 2: Price Changes ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Price Changes")
    ws2.freeze_panes = "A3"
    ws2.merge_cells("A1:K1")
    c2 = ws2.cell(row=1, column=1, value="Price Changes Detected vs Baseline (~1 year ago)")
    c2.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    c2.fill = PatternFill("solid", start_color="856404")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 22

    pc_headers = ["branch_id","restaurant_name","item_name","category",
                  "old_price_aed","new_price_aed","change_pct","scraped_at"]
    pc_widths  = [12, 32, 38, 24, 14, 14, 13, 22]
    for col, (h, w) in enumerate(zip(pc_headers, pc_widths), 1):
        _cell(ws2, 2, col, h, bold=True, bg=HDR_BG, fg="FFFFFF", align="center")
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[2].height = 18

    price_rows = [r for r in all_deltas if r["change_type"] == "CHANGED" and r.get("field_changed") == "price"]
    for i, r in enumerate(price_rows, 3):
        pct = r.get("price_change_pct")
        pct_fg = "155724" if (pct or 0) < 0 else ("721C24" if (pct or 0) > 0 else "000000")
        bg = "FFF3CD" if i % 2 == 0 else "FFFDE7"
        for col, (key, fmt) in enumerate([
            ("branch_id", None), ("restaurant_name", None), ("item_name", None),
            ("category", None), ("old_price_aed", "#,##0.00"),
            ("new_price_aed", "#,##0.00"), ("price_change_pct", "+0.00;-0.00"),
            ("scraped_at", None),
        ], 1):
            cc = _cell(ws2, i, col, r.get(key), bg=bg,
                       align="center" if col in (1,5,6,7) else "left",
                       num_fmt=fmt)
            if col == 7:
                cc.font = Font(name="Arial", bold=True, color=pct_fg, size=9)
        ws2.row_dimensions[i].height = 15

    # ── Sheet 3: New Items ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("New Items")
    ws3.freeze_panes = "A3"
    ws3.merge_cells("A1:H1")
    c3 = ws3.cell(row=1, column=1, value="Items Added Since Baseline")
    c3.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    c3.fill = PatternFill("solid", start_color="155724")
    c3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 22

    ni_headers = ["branch_id","restaurant_name","item_name","category","new_price_aed","description","scraped_at"]
    ni_widths  = [12, 32, 38, 24, 14, 50, 22]
    for col, (h, w) in enumerate(zip(ni_headers, ni_widths), 1):
        _cell(ws3, 2, col, h, bold=True, bg=HDR_BG, fg="FFFFFF", align="center")
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.row_dimensions[2].height = 18

    added_rows = [r for r in all_deltas if r["change_type"] == "ADDED"]
    for i, r in enumerate(added_rows, 3):
        bg = "D4EDDA" if i % 2 == 0 else "EBF7EE"
        for col, key in enumerate(["branch_id","restaurant_name","item_name",
                                   "category","new_price_aed","description","scraped_at"], 1):
            _cell(ws3, i, col, r.get(key), bg=bg,
                  align="center" if col in (1,5) else "left",
                  num_fmt="#,##0.00" if col == 5 else None,
                  wrap=col == 6)
        ws3.row_dimensions[i].height = 15

    # ── Sheet 4: Removed Items ───────────────────────────────────────────────
    ws4 = wb.create_sheet("Removed Items")
    ws4.freeze_panes = "A3"
    ws4.merge_cells("A1:H1")
    c4 = ws4.cell(row=1, column=1, value="Items Removed Since Baseline")
    c4.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    c4.fill = PatternFill("solid", start_color="721C24")
    c4.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 22

    rm_headers = ["branch_id","restaurant_name","item_name","category","old_price_aed","description","scraped_at"]
    rm_widths  = [12, 32, 38, 24, 14, 50, 22]
    for col, (h, w) in enumerate(zip(rm_headers, rm_widths), 1):
        _cell(ws4, 2, col, h, bold=True, bg=HDR_BG, fg="FFFFFF", align="center")
        ws4.column_dimensions[get_column_letter(col)].width = w
    ws4.row_dimensions[2].height = 18

    removed_rows = [r for r in all_deltas if r["change_type"] == "REMOVED"]
    for i, r in enumerate(removed_rows, 3):
        bg = "F8D7DA" if i % 2 == 0 else "FDECEE"
        for col, key in enumerate(["branch_id","restaurant_name","item_name",
                                   "category","old_price_aed","description","scraped_at"], 1):
            _cell(ws4, i, col, r.get(key), bg=bg,
                  align="center" if col in (1,5) else "left",
                  num_fmt="#,##0.00" if col == 5 else None,
                  wrap=col == 6)
        ws4.row_dimensions[i].height = 15

    # ── Sheet 5: Full Current Menu ───────────────────────────────────────────
    ws5 = wb.create_sheet("Full Menu Current")
    ws5.freeze_panes = "A3"
    ws5.merge_cells("A1:J1")
    c5 = ws5.cell(row=1, column=1,
                  value=f"Full Menu — All Scraped Items | Run {run_id}")
    c5.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    c5.fill = PatternFill("solid", start_color="2E4057")
    c5.alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 22

    fm_headers = ["branch_id","restaurant_name","area_name","category",
                  "item_name","price_aed","description","image_url","item_id","scraped_at"]
    fm_widths  = [12, 32, 20, 26, 38, 12, 50, 40, 16, 22]
    for col, (h, w) in enumerate(zip(fm_headers, fm_widths), 1):
        _cell(ws5, 2, col, h, bold=True, bg=HDR_BG, fg="FFFFFF", align="center")
        ws5.column_dimensions[get_column_letter(col)].width = w
    ws5.row_dimensions[2].height = 18

    row5 = 3
    for snap in all_snapshots:
        if snap.get("status") not in ("ok", "changed", "no_change", "new", "new_restaurant"):
            continue
        for item in snap.get("items", []):
            bg = "F5F5F5" if row5 % 2 == 0 else "FFFFFF"
            vals = [
                snap.get("branch_id"), snap.get("restaurant_name",""),
                snap.get("area_name",""), item.get("category",""),
                item.get("item_name",""), item.get("price_aed"),
                item.get("description",""), item.get("image_url",""),
                item.get("item_id"), snap.get("scraped_at",""),
            ]
            for col, (val, fmt) in enumerate(
                zip(vals, [None,None,None,None,None,"#,##0.00",None,None,None,None]), 1
            ):
                _cell(ws5, row5, col, val, bg=bg,
                      align="center" if col in (1,6,9) else "left",
                      num_fmt=fmt, wrap=col == 7)
            ws5.row_dimensions[row5].height = 15
            row5 += 1

    out_path = REPORT_DIR / f"menu_delta_report_{run_id}.xlsx"
    wb.save(out_path)
    print(f"Excel report : {out_path}")
    return out_path


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Talabat Menu Tracker")
    parser.add_argument("--all",     action="store_true", help="Run all 5,275 restaurants")
    parser.add_argument("--rebuild", action="store_true", help="Force rebuild baseline from Excel")
    parser.add_argument("--limit",   type=int, default=TEST_LIMIT, help="URL limit (default 200)")
    args = parser.parse_args()

    run_mode  = "FULL" if args.all else f"TEST ({args.limit})"
    limit     = None if args.all else args.limit
    run_id    = datetime.now().strftime("%Y%m%d_%H%M%S")

    print("=" * 70)
    print(f"TALABAT MENU TRACKER — Run {run_id}")
    print(f"Mode     : {run_mode}")
    print(f"Workers  : {CONCURRENT_WORKERS}")
    print(f"Delay    : {MIN_DELAY}–{MAX_DELAY}s")
    print(f"Output   : {DATA_DIR}")
    print("=" * 70)

    if not USERNAME or not PASSWORD:
        print("ERROR: Oxylabs credentials not set in talabat/.env")
        sys.exit(1)

    # 1. Build / load baseline
    baseline = build_baseline(force=args.rebuild)
    print(f"Baseline: {len(baseline)} restaurants loaded")

    # 2. Resume
    done = load_checkpoint()
    processed_ids.update(done)
    if done:
        print(f"Resuming: {len(done)} already done")

    # 3. Build job list: unique (branch_id, url) pairs
    all_jobs = [(bid, bdata["url"]) for bid, bdata in baseline.items()
                if bid not in done]
    if limit:
        all_jobs = all_jobs[:limit]

    print(f"Jobs to process: {len(all_jobs)}")
    if not all_jobs:
        print("All done. Generating outputs from existing data...")
    else:
        # Estimate
        avg_d = (MIN_DELAY + MAX_DELAY) / 2
        est   = (len(all_jobs) * avg_d) / CONCURRENT_WORKERS
        print(f"Est. time: ~{est/60:.0f} min")
        print()

        # 4. Split & dispatch
        n = CONCURRENT_WORKERS
        batch_sz = max(1, -(-len(all_jobs) // n))
        batches  = [all_jobs[i:i+batch_sz] for i in range(0, len(all_jobs), batch_sz)]
        while len(batches) > n:
            batches[-2].extend(batches.pop())
        print(f"Batch sizes: {[len(b) for b in batches]}")
        print()

        with ThreadPoolExecutor(max_workers=n) as ex:
            futs = {ex.submit(worker, i+1, batch, baseline): i for i, batch in enumerate(batches)}
            for fut in as_completed(futs):
                try:
                    fut.result()
                except Exception as e:
                    print(f"Worker {futs[fut]+1} raised: {e}")

    # 5. Final checkpoint + outputs
    save_checkpoint(processed_ids)
    write_json(run_id)
    write_excel(run_id)

    # 6. Summary
    print()
    print("=" * 70)
    print("RUN COMPLETE")
    print(f"  Scraped          : {counters['scraped']:,}")
    print(f"  Changed (delta)  : {counters['changed']:,}")
    print(f"  No change        : {counters['no_change']:,}")
    print(f"  New (no baseline): {counters['new']:,}")
    print(f"  Failed           : {counters['failed']:,}")
    print(f"  Price changes    : {len([r for r in all_deltas if r['change_type']=='CHANGED' and r.get('field_changed')=='price']):,}")
    print(f"  Items added      : {len([r for r in all_deltas if r['change_type']=='ADDED']):,}")
    print(f"  Items removed    : {len([r for r in all_deltas if r['change_type']=='REMOVED']):,}")
    print("=" * 70)


if __name__ == "__main__":
    main()
